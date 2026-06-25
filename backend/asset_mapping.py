"""
Asset classification loader — single source of truth is data/master/Asset_Master.xlsx.

Public API
----------
load_asset_mapping(data_dir)   → mapping dict (cached by file mtime+size)
classify_work_order(record, mapping) → classification dict
build_refrigeration_tree(mapping)    → list of subgroup dicts

The mapping dict has keys:
  available, path, last_synced, asset_map, keyword_rules,
  groups, group_matchers, message
"""

import os
import re
from datetime import datetime
from pathlib import Path

import openpyxl

ASSET_MASTER_FILENAME = "Asset_Master.xlsx"
ASSET_MASTER_RELATIVE_PATH = Path("master") / ASSET_MASTER_FILENAME
ASSET_MAPPING_SHEET = "Asset_Master"
KEYWORD_RULES_SHEET = "Keyword Rules"

MACHINE_GROUPS = [
    "Refrigeration",
    "Production Equipment",
    "Utilities / Support",
    "Facility / Building",
    "Unknown / Review",
]

CRITICALITIES = ["Critical", "Non-Critical", "Facility", "Unmapped"]

# Machine groups that are treated as Critical (operational / production-critical).
# Any asset whose Main Asset Group falls in this set is labelled Critical so that
# the MTTR / MTBF dashboards filter correctly when the user selects "Critical".
_CRITICAL_MACHINE_GROUPS = {"Production Equipment", "Utilities", "Utilities / Support", "Refrigeration"}

REFRIGERATION_ROLES = [
    "Condenser", "Evaporator", "Freezer", "Chiller",
    "Cold Room", "Ice Maker", "Other",
]

REFRIGERATION_SUBGROUPS = [
    "condenser-evaporator", "air-blast", "cold-room", "ice-maker", "other",
]

_MACHINE_GROUP_SET = set(MACHINE_GROUPS)

_CANDIDATES = [
    Path(__file__).resolve().parents[1] / "data" / ASSET_MASTER_RELATIVE_PATH,
    Path(__file__).resolve().parents[1] / "data" / ASSET_MASTER_FILENAME,
    Path(__file__).resolve().parent / ASSET_MASTER_FILENAME,
    Path.home() / "Downloads" / ASSET_MASTER_FILENAME,
]

_CACHE = {"sig": None, "payload": None}


def _file_sig(path):
    try:
        s = os.stat(path)
        return (s.st_mtime_ns, s.st_size)
    except OSError:
        return None


def _clean(value, fallback=""):
    text = re.sub(r"\s+", " ", str(value or "").replace("﻿", " ").strip())
    return text or fallback


def _is_banner(value):
    return str(value or "").lstrip().startswith("▶")


def _truthy(value):
    if value is None:
        return True
    return str(value).strip().upper() not in {"FALSE", "0", "NO", "N"}


def _normalize_criticality(value):
    k = re.sub(r"[^a-z]", "", str(value or "").lower())
    if k in {"critical", "semicritical", "productioncritical"}:
        return "Critical"
    if k in {"noncritical", "noncriticalfacility", "facility"}:
        return "Non-Critical"
    if k == "facility":
        return "Facility"
    return "Unmapped"


def _normalize_machine_name(value):
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _normalize_header(value):
    return re.sub(r"[^a-z0-9]", "", str(value or "").strip().lower())


def _normalize_stage(value):
    text = _clean(value)
    normalized = re.sub(r"\s+", "", text.lower())
    if normalized in {"stage1", "st1", "s1"}:
        return "Stage 1"
    if normalized in {"stage2", "st2", "s2"}:
        return "Stage 2"
    return "Needs Stage Review"


def _stage_mapping_status(stage, asset_id):
    if not asset_id:
        return "Missing Asset ID"
    return "Mapped" if stage in {"Stage 1", "Stage 2"} else "Needs Stage Review"


def _build_group_aliases(name):
    aliases = {_normalize_machine_name(name), re.sub(r"[^a-z0-9]", "", _normalize_machine_name(name))}
    return sorted({a for a in aliases if a}, key=len, reverse=True)


def _resolve_path(data_dir):
    candidates = [Path(data_dir) / ASSET_MASTER_RELATIVE_PATH, Path(data_dir) / ASSET_MASTER_FILENAME] + list(_CANDIDATES)
    seen = set()
    for p in candidates:
        try:
            r = p.resolve()
        except OSError:
            r = p
        if r in seen:
            continue
        seen.add(r)
        sig = _file_sig(p)
        if sig:
            return p, sig
    return None, None


def load_asset_mapping(data_dir):
    path, sig = _resolve_path(data_dir)

    if sig and _CACHE["sig"] == sig and _CACHE["payload"] is not None:
        return _CACHE["payload"]

    empty = {
        "available": False,
        "path": str(path or (Path(data_dir) / ASSET_MASTER_RELATIVE_PATH)),
        "last_synced": None,
        "asset_map": {},
        "keyword_rules": [],
        "groups": [],
        "group_matchers": [],
        "message": f"{ASSET_MASTER_RELATIVE_PATH.as_posix()} not found.",
    }

    if not path:
        _CACHE.update(sig=None, payload=empty)
        return empty

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        empty["message"] = f"Cannot open {ASSET_MASTER_FILENAME}: {exc}"
        raise RuntimeError(empty["message"]) from exc

    if ASSET_MAPPING_SHEET not in wb.sheetnames:
        msg = f"{ASSET_MASTER_FILENAME} is missing the '{ASSET_MAPPING_SHEET}' sheet."
        raise RuntimeError(msg)

    # ── Parse Asset Mapping ───────────────────────────────────────────────────
    ws = wb[ASSET_MAPPING_SHEET]
    asset_map = {}
    groups = []
    group_matchers = []

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
    header_index = {_normalize_header(value): idx for idx, value in enumerate(header_row)}

    def cell(row, *names, fallback_index=None):
        for name in names:
            idx = header_index.get(_normalize_header(name))
            if idx is not None and idx < len(row):
                return row[idx]
        if fallback_index is not None and fallback_index < len(row):
            return row[fallback_index]
        return None

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_id = cell(row, "Asset ID", "AssetID", fallback_index=0)
        if raw_id is None:
            continue
        if _is_banner(raw_id):
            continue
        asset_id = _clean(raw_id).upper()
        if not asset_id:
            continue

        display_name = _clean(cell(row, "Asset Name", "DisplayName", fallback_index=1), asset_id)
        stage = _normalize_stage(cell(row, "Stage", fallback_index=2))
        mapping_status = _stage_mapping_status(stage, asset_id)
        machine_group = _clean(cell(row, "Category", "Main Asset Group", fallback_index=3), "Unknown / Review")
        asset_machine_group = _clean(cell(row, "Machine Group"), "")
        sub_asset_group = _clean(cell(row, "Sub Asset Group", fallback_index=5), "")
        location = _clean(cell(row, "Location", fallback_index=5), "Unassigned")
        system_area = _clean(cell(row, "System/Area", "System Area", fallback_index=6), "")
        remarks = _clean(cell(row, "Remarks", fallback_index=7), "")

        raw_criticality = ""
        if mapping_status != "Mapped":
            criticality = "Unmapped"
            criticality_rank = 999
        elif machine_group in _CRITICAL_MACHINE_GROUPS:
            # Production Equipment, Utilities, and Refrigeration are operational-
            # critical assets; they must appear under the "Critical" filter in the
            # MTTR / MTBF sections.
            criticality = "Critical"
            criticality_rank = 1
        else:
            criticality = "Non-Critical"
            criticality_rank = 2
        refrigeration_role = "Other" if machine_group == "Refrigeration" else ""
        parent_asset_id = None
        refrigeration_subgroup = sub_asset_group or None

        entry = {
            "asset_id": asset_id,
            "display_name": display_name,
            "machine_group": machine_group,
            "asset_machine_group": asset_machine_group,
            "location": location,
            "stage": stage,
            "mappedStage": stage,
            "mappedAssetName": display_name,
            "mappedMainAssetGroup": machine_group,
            "mappedMachineGroup": asset_machine_group,
            "mappedSubAssetGroup": sub_asset_group,
            "mappedLocation": location,
            "mappedSystemArea": system_area,
            "mappingStatus": mapping_status,
            "mapped_stage": stage,
            "mapped_asset_name": display_name,
            "mapped_main_asset_group": machine_group,
            "mapped_sub_asset_group": sub_asset_group,
            "mapped_location": location,
            "mapped_system_area": system_area,
            "mapping_status": mapping_status,
            "remarks": remarks,
            "criticality": criticality,
            "raw_criticality": raw_criticality,
            "criticality_rank": criticality_rank,
            "refrigeration_role": refrigeration_role or None,
            "parent_asset_id": parent_asset_id,
            "refrigeration_subgroup": refrigeration_subgroup or None,
            "mapping_source": ASSET_MASTER_FILENAME,
            "classification_source": ASSET_MASTER_FILENAME,
            "has_assetlist_classification": mapping_status == "Mapped",
            "has_asset_master_mapping": mapping_status == "Mapped",
            # Legacy compat fields used by downtime_management
            "machine_name_display": display_name,
            "asset_label": asset_id,
            "asset_display_name": display_name,
            "building": location,
            "group_asset_ids": [],
        }
        asset_map[asset_id] = entry

        machine_name_lc = _normalize_machine_name(machine_group)
        group_key = machine_group
        existing = next((g for g in groups if g["machine_group"] == group_key), None)
        if not existing:
            g = {
                "machine_group": machine_group,
                "machine_name_display": machine_group,
                "location": location,
                "building": location,
                "criticality": criticality,
                "raw_criticality": raw_criticality,
                "criticality_rank": criticality_rank,
                "classification_source": ASSET_MASTER_FILENAME,
                "has_assetlist_classification": mapping_status == "Mapped",
                "mappedStage": stage,
                "mappedMainAssetGroup": machine_group,
                "mappedSubAssetGroup": sub_asset_group,
                "mappedLocation": location,
                "mappedSystemArea": system_area,
                "mappingStatus": mapping_status,
                "asset_ids": [],
                "asset_entries": [],
            }
            groups.append(g)
            group_matchers.append({
                "machine_group": machine_group,
                "machine_name_display": machine_group,
                "location": location,
                "building": location,
                "criticality": criticality,
                "raw_criticality": raw_criticality,
                "criticality_rank": criticality_rank,
                "classification_source": ASSET_MASTER_FILENAME,
                "has_assetlist_classification": mapping_status == "Mapped",
                "mappedStage": stage,
                "mappedMainAssetGroup": machine_group,
                "mappedSubAssetGroup": sub_asset_group,
                "mappedLocation": location,
                "mappedSystemArea": system_area,
                "mappingStatus": mapping_status,
                "aliases": _build_group_aliases(machine_group),
            })
            existing = g
        existing["asset_ids"].append(asset_id)
        existing["asset_entries"].append({
            "asset_id": asset_id,
            "asset_label": asset_id,
            "asset_display_name": display_name,
            "mappedStage": stage,
            "mappedAssetName": display_name,
            "mappedMainAssetGroup": machine_group,
            "mappedMachineGroup": asset_machine_group,
            "mappedSubAssetGroup": sub_asset_group,
            "mappedLocation": location,
            "mappedSystemArea": system_area,
            "mappingStatus": mapping_status,
        })

    # Back-fill group_asset_ids into each asset_map entry
    for group in groups:
        for asset_id in group["asset_ids"]:
            if asset_id in asset_map:
                asset_map[asset_id]["group_asset_ids"] = group["asset_ids"]

    # ── Parse Keyword Rules ───────────────────────────────────────────────────
    keyword_rules = []
    if KEYWORD_RULES_SHEET in wb.sheetnames:
        ws2 = wb[KEYWORD_RULES_SHEET]
        for row in ws2.iter_rows(min_row=3, values_only=True):
            kw = row[0] if row else None
            if kw is None:
                continue
            if _is_banner(kw):
                continue
            keyword = _clean(kw).lower()
            maps_to = _clean(row[1] if len(row) > 1 else None)
            active = _truthy(row[2] if len(row) > 2 else None)
            if not keyword or not maps_to or not active:
                continue
            keyword_rules.append({"keyword": keyword, "maps_to": maps_to})

    # Skip header row if it leaked through
    keyword_rules = [r for r in keyword_rules if r["keyword"] != "keyword"]

    wb.close()

    payload = {
        "available": True,
        "path": str(path),
        "last_synced": datetime.fromtimestamp(os.path.getmtime(path)).isoformat(),
        "asset_map": asset_map,
        "keyword_rules": keyword_rules,
        "groups": groups,
        "group_matchers": sorted(
            group_matchers,
            key=lambda m: max((len(a) for a in m["aliases"]), default=0),
            reverse=True,
        ),
        "message": f"{ASSET_MASTER_FILENAME} loaded ({len(asset_map)} assets, {len(keyword_rules)} keyword rules).",
    }
    _CACHE.update(sig=sig, payload=payload)
    return payload


# ── Shared Production-Equipment / Utilities / Unclassified classifier ─────────
# Maps a Main Asset Group (from Asset_Master) to the two operational categories the
# Downtime and Spare-Parts pages filter by. Anything that is not clearly production
# or utility/support is Unclassified (per the brief — never guess-map).
_GROUP_TO_CATEGORY = {
    "Production Equipment": "Production Equipment",
    "Utilities / Support": "Utilities",
    "Utilities": "Utilities",
    "Refrigeration": "Utilities",          # cooling/refrigeration = support utility
    "Facility / Building": "Unclassified",  # facility (lighting/doors/CCTV) — neither
    "Unknown / Review": "Unclassified",
}
ASSET_CATEGORIES = ["Production Equipment", "Utilities", "Unclassified"]


def group_to_category(machine_group: str) -> str:
    return _GROUP_TO_CATEGORY.get(str(machine_group or "").strip(), "Unclassified")


def classify_asset_category(asset_id=None, asset_name=None, functional_location=None, mapping=None):
    """Classify an asset/item into Production Equipment | Utilities | Unclassified.

    Asset Master first (by asset_id); keyword fallback (asset_name / functional
    location) only when the asset_id isn't mapped — matching the brief's rule
    "first use the Asset Master; keyword fallback only if asset master is missing".
    Returns "Unclassified" for anything that can't be mapped to a real category."""
    if mapping is None:
        try:
            mapping = load_asset_mapping(str(Path(__file__).resolve().parents[1] / "data"))
        except Exception:
            return "Unclassified"
    record = {
        "asset_id": asset_id or "",
        "machine_name": asset_name or "",
        "raw_machine_name": asset_name or "",
        "raw_functional_location": functional_location or "",
    }
    try:
        result = classify_work_order(record, mapping)
    except Exception:
        return "Unclassified"
    group = result.get("machine_group") or result.get("mappedMainAssetGroup") or ""
    return group_to_category(group)


def classify_work_order(record, mapping):
    """
    Classify a work order record against the asset mapping.

    Returns a dict with: asset_id, display_name, machine_group, criticality,
    raw_criticality, criticality_rank, refrigeration_role, parent_asset_id,
    refrigeration_subgroup, mapping_source, classification_source,
    has_assetlist_classification, mappedStage, mappedAssetName,
    mappedMainAssetGroup, mappedSubAssetGroup, mappedLocation,
    mappedSystemArea, mappingStatus, group_asset_ids,
    machine_name_display, asset_label, asset_display_name, location, building.
    """
    asset_map = mapping.get("asset_map", {})
    keyword_rules = mapping.get("keyword_rules", [])

    raw_id = str(record.get("asset_id") or record.get("machine_code") or "").strip().upper()

    # Step A — direct match
    hit = asset_map.get(raw_id)
    if hit:
        return dict(hit)

    # Step B — keyword fallback
    searchable = " ".join(
        str(record.get(field) or "")
        for field in (
            "asset_id", "machine_code", "machine_name", "raw_machine_name",
            "machine_equipment_name", "description", "description_original",
            "remarks", "job_trade", "system", "maintenance_job_type",
            "raw_functional_location", "area", "location",
        )
    ).lower()

    for rule in keyword_rules:
        if rule["keyword"] in searchable:
            maps_to = rule["maps_to"]
            if maps_to in _MACHINE_GROUP_SET:
                return _fallback_entry(raw_id, record, maps_to, "keyword")
            # Treat MapsTo as a DisplayName → find any matching asset row
            display_hit = next(
                (e for e in asset_map.values() if e["display_name"] == maps_to),
                None,
            )
            if display_hit:
                result = dict(display_hit)
                result["display_name"] = maps_to
                result["machine_name_display"] = maps_to
                result["mapping_source"] = "keyword"
                result["classification_source"] = "keyword"
                result["mappingStatus"] = "Keyword Matched"
                result["mapping_status"] = "Keyword Matched"
                result["has_assetlist_classification"] = False
                result["has_asset_master_mapping"] = False
                return result
            # MapsTo doesn't match a display name — treat as a machine-group label
            return _fallback_entry(raw_id, record, maps_to, "keyword")

    # Step C — unmapped
    return _fallback_entry(raw_id, record, "Unknown / Review", "fallback")


def _fallback_entry(asset_id, record, machine_group, source):
    display = (
        str(record.get("raw_machine_name") or record.get("machine_name") or "").strip()
        or str(record.get("machine_equipment_name") or "").strip()
        or asset_id
        or "Unmapped Asset"
    )
    location = str(record.get("raw_functional_location") or record.get("area") or record.get("location") or "Unassigned").strip()
    is_missing_asset_id = not str(asset_id or "").strip()
    if source == "keyword":
        mapping_status = "Keyword Matched"
        mapped_stage = "Needs Stage Review"
    elif is_missing_asset_id:
        mapping_status = "Missing Asset ID"
        mapped_stage = "Missing Asset ID"
    else:
        mapping_status = "Unmapped"
        mapped_stage = "Unmapped"
    criticality = "Unmapped" if machine_group in {"Unknown / Review", "Unmapped"} else "Non-Critical"
    criticality_rank = 999 if criticality == "Unmapped" else 2
    return {
        "asset_id": asset_id,
        "display_name": display,
        "machine_group": machine_group,
        "location": location,
        "building": location,
        "stage": mapped_stage,
        "mappedStage": mapped_stage,
        "mappedAssetName": display,
        "mappedMainAssetGroup": machine_group,
        "mappedSubAssetGroup": "",
        "mappedLocation": location,
        "mappedSystemArea": "",
        "mappingStatus": mapping_status,
        "mapped_stage": mapped_stage,
        "mapped_asset_name": display,
        "mapped_main_asset_group": machine_group,
        "mapped_sub_asset_group": "",
        "mapped_location": location,
        "mapped_system_area": "",
        "mapping_status": mapping_status,
        "criticality": criticality,
        "raw_criticality": "",
        "criticality_rank": criticality_rank,
        "refrigeration_role": None,
        "parent_asset_id": None,
        "refrigeration_subgroup": None,
        "mapping_source": source,
        "classification_source": source,
        "has_assetlist_classification": False,
        "has_asset_master_mapping": False,
        "machine_name_display": display,
        "asset_label": asset_id,
        "asset_display_name": display,
        "group_asset_ids": [asset_id] if asset_id else [],
    }


def build_refrigeration_tree(mapping):
    """
    Return a list of subgroup dicts, each containing parent assets (Condensers)
    with their children (Evaporators) nested inside.

    [
      { subgroup: "condenser-evaporator", parents: [
          { asset_id, display_name, role, children: [ ... ] }
        ]
      },
      ...
    ]
    """
    asset_map = mapping.get("asset_map", {})
    subgroups = {}

    for asset_id, entry in asset_map.items():
        if entry.get("machine_group") != "Refrigeration":
            continue
        sg = entry.get("refrigeration_subgroup") or "other"
        subgroups.setdefault(sg, {"parents": {}, "children": []})

        if not entry.get("parent_asset_id"):
            # It's a parent (Condenser or standalone unit)
            subgroups[sg]["parents"].setdefault(asset_id, {
                "asset_id": asset_id,
                "display_name": entry["display_name"],
                "refrigeration_role": entry.get("refrigeration_role"),
                "criticality": entry["criticality"],
                "location": entry["location"],
                "children": [],
            })
        else:
            subgroups[sg]["children"].append(entry)

    # Nest children under their parents
    for sg_data in subgroups.values():
        for child in sg_data["children"]:
            parent_id = child["parent_asset_id"]
            # Determine subgroup from parent if parent exists
            for sg_key, data in subgroups.items():
                if parent_id in data["parents"]:
                    data["parents"][parent_id]["children"].append({
                        "asset_id": child["asset_id"],
                        "display_name": child["display_name"],
                        "refrigeration_role": child.get("refrigeration_role"),
                        "criticality": child["criticality"],
                        "parent_asset_id": parent_id,
                    })
                    break

    result = []
    for sg_key in REFRIGERATION_SUBGROUPS + ["other"]:
        if sg_key not in subgroups:
            continue
        parents_sorted = sorted(subgroups[sg_key]["parents"].values(), key=lambda p: p["asset_id"])
        for parent in parents_sorted:
            parent["children"].sort(key=lambda c: c["asset_id"])
        result.append({"subgroup": sg_key, "parents": parents_sorted})

    return result


def get_asset_mapping_meta(data_dir):
    m = load_asset_mapping(data_dir)
    return {
        "available": m["available"],
        "path": m["path"],
        "last_synced": m["last_synced"],
        "asset_count": len(m["asset_map"]),
        "keyword_rule_count": len(m["keyword_rules"]),
        "group_count": len(m["groups"]),
        "message": m["message"],
    }
