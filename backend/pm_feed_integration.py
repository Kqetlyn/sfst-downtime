"""
D365 PM feed integration (Stage 2 PM tasks).

Replaces the old hard-coded / week-token Stage 2 generator with the live D365 PM
feed. The feed arrives as two workbooks in the same shape (production + utility),
each with two sheets:

    5_MaintPlans:     Plan ID*, Name*, "Work order type*", Trade*, Active*, Description
    6_MaintPlanLines: Plan ID*, "Asset or Functional location*", "Reference type*",
                      "Reference value*", "Interval unit*", "Interval value*",
                      "Start date*", "Job type", "Job description"

Classification is driven ENTIRELY by Asset_Master.xlsx (never hard-coded):
    - Asset_Master sheet:  Asset ID -> Stage, Main Asset Group, Sub Asset Group,
                           Location, System/Area, "PIC / Owner"
    - PM_Feed_Map sheet:   Feed Code -> Stage, Scope, System/Area, Location, Default PIC
    - Lists sheet D/E:     Main Asset Group -> Scope

Each plan line is resolved MOST-SPECIFIC-FIRST:
    a) the line names an Asset ID (col "Asset or Functional location*" matches
       ``^EN[A-Z]{2}-`` OR Reference type = Asset) -> Asset_Master lookup; Scope is
       derived from Main Asset Group via the Lists D/E map.
    b) otherwise the "Reference value*" code -> PM_Feed_Map lookup.
    c) any field still blank is filled from whichever source the other path didn't.

This ordering is what makes the Spiral Freezers (referenced by a packing-line code
but whose ASSET is Refrigeration) classify as Utilities scope.

Recurrence: from "Start date*", step by ("Interval value*" x "Interval unit*")
to emit one task per occurrence inside the window. Plans with Active* = No are
skipped.

This module is pure/standalone (openpyxl only) so it can be unit-checked against
the real workbooks before it is wired into ``pm_schedule_service``.
"""

from __future__ import annotations

import calendar
import re
from datetime import date, datetime
from pathlib import Path

import openpyxl

ASSET_ID_RE = re.compile(r"^EN[A-Z]{2}-", re.IGNORECASE)

# Feed source descriptors (the two workbooks).
PRODUCTION_FEED = "production"
UTILITY_FEED = "utility"

# The D365 PM feed is the Stage 2 PM source. Stage is read from the master
# (Asset_Master / PM_Feed_Map, which are all Stage 2); this is only the safety
# fallback used when a resolved line leaves Stage blank, so a feed task is never
# stage-less. It is NOT a hard override of workbook values.
DEFAULT_FEED_STAGE = "Stage 2"

# Canonical filenames in DATA_DIR (swap data, not code: a new machine/code is an
# Excel edit to the feed + Asset_Master, never a code change).
PRODUCTION_FEED_FILENAME = "PM_template_filled production machines.xlsx"
UTILITY_FEED_FILENAME = "PM_template_filled v1.xlsx"
MASTER_FILENAME = "Asset_Master.xlsx"


def default_feeds(data_dir) -> list[dict]:
    base = Path(data_dir)
    return [
        {"path": base / PRODUCTION_FEED_FILENAME, "source": PRODUCTION_FEED},
        {"path": base / UTILITY_FEED_FILENAME, "source": UTILITY_FEED},
    ]


def default_master_path(data_dir) -> Path:
    return Path(data_dir) / "master" / MASTER_FILENAME

_INTERVAL_DAYS = {"day": 1, "days": 1, "week": 7, "weeks": 7}
_SAFETY_CAP = 4000  # max occurrences expanded per line (guards bad interval data)


# ── small helpers ────────────────────────────────────────────────────────────
def _txt(value) -> str:
    return "" if value is None else str(value).strip()


def _norm_header(value) -> str:
    """Normalise a header cell to a comparison key: lowercase, alnum only."""
    return re.sub(r"[^a-z0-9]", "", _txt(value).lower())


def _coerce_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text[:10]).date()
    except ValueError:
        return None


def _is_active(value) -> bool:
    """Active* truthy unless explicitly No/False/0/Inactive."""
    return _txt(value).lower() not in {"no", "false", "0", "inactive", "n"}


def _header_index(header_row, *aliases) -> int | None:
    """Find the column index whose header matches any alias (fuzzy)."""
    keys = [_norm_header(a) for a in aliases]
    norm = [_norm_header(c) for c in header_row]
    # exact-normalised match first
    for i, h in enumerate(norm):
        if h in keys:
            return i
    # then contains match (handles "Plan ID* (specify what this is)" etc.)
    for i, h in enumerate(norm):
        if h and any(k and (k in h or h in k) for k in keys):
            return i
    return None


def _find_header_row(rows, *required_aliases, limit=4):
    """Return (row_index, header_list) for the first row that contains all
    required aliases. Handles sheets where row 1 is a note and row 2 is the
    header (e.g. PM_Feed_Map)."""
    for idx in range(min(limit, len(rows))):
        header = list(rows[idx])
        if all(_header_index(header, alias) is not None for alias in required_aliases):
            return idx, header
    return None, None


# ── recurrence ───────────────────────────────────────────────────────────────
def _add_interval(start: date, unit: str, value: int) -> date:
    unit = (unit or "").strip().lower()
    value = int(value)
    if unit in ("month", "months"):
        total = (start.month - 1) + value
        year = start.year + total // 12
        month = total % 12 + 1
        day = min(start.day, calendar.monthrange(year, month)[1])
        return date(year, month, day)
    if unit in ("year", "years"):
        year = start.year + value
        day = min(start.day, calendar.monthrange(year, start.month)[1])
        return date(year, start.month, day)
    days = _INTERVAL_DAYS.get(unit)
    if days:
        return date.fromordinal(start.ordinal() if False else start.toordinal() + days * value)
    # Unknown unit -> treat as a single (non-recurring) occurrence.
    return None


def expand_occurrences(start: date, unit: str, value, win_start: date, win_end: date) -> list[date]:
    """One date per occurrence of (start stepped by value x unit) within window."""
    if not start:
        return []
    try:
        step = int(value)
    except (TypeError, ValueError):
        step = 0
    out: list[date] = []
    if step <= 0:
        # No valid interval -> a single occurrence at the start date.
        if win_start <= start <= win_end:
            out.append(start)
        return out
    current = start
    guard = 0
    while current <= win_end and guard < _SAFETY_CAP:
        if current >= win_start:
            out.append(current)
        nxt = _add_interval(current, unit, step)
        if nxt is None or nxt <= current:
            break
        current = nxt
        guard += 1
    return out


# ── feed reading ─────────────────────────────────────────────────────────────
def read_feed(path: str | Path, source: str) -> dict:
    """Parse one feed workbook into joined plan-line records.

    Returns {"lines": [record, ...], "errors": [...]}; each record carries the
    plan-level fields it needs (active, name, work order type, trade) plus the
    line fields.
    """
    path = Path(path)
    result = {"lines": [], "errors": [], "source": source}
    if not path.exists():
        result["errors"].append(f"Feed workbook not found: {path}")
        return result
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        plans = _read_plans(wb, result)
        result["lines"] = _read_plan_lines(wb, plans, source, result)
    finally:
        wb.close()
    return result


def _read_plans(wb, result) -> dict:
    if "5_MaintPlans" not in wb.sheetnames:
        result["errors"].append("Missing sheet 5_MaintPlans")
        return {}
    rows = list(wb["5_MaintPlans"].iter_rows(values_only=True))
    hdr_idx, header = _find_header_row(rows, "Plan ID", "Active")
    if header is None:
        result["errors"].append("5_MaintPlans: header row not found")
        return {}
    ci = {
        "plan": _header_index(header, "Plan ID"),
        "name": _header_index(header, "Name"),
        "wo_type": _header_index(header, "Work order type"),
        "trade": _header_index(header, "Trade"),
        "active": _header_index(header, "Active"),
        "desc": _header_index(header, "Description"),
    }
    plans = {}
    for row in rows[hdr_idx + 1:]:
        plan_id = _txt(row[ci["plan"]]) if ci["plan"] is not None and ci["plan"] < len(row) else ""
        if not plan_id:
            continue
        def cell(key):
            i = ci[key]
            return _txt(row[i]) if i is not None and i < len(row) else ""
        plans[plan_id] = {
            "name": cell("name"),
            "workOrderType": cell("wo_type"),
            "trade": cell("trade"),
            "active": _is_active(cell("active")),
            "description": cell("desc"),
        }
    return plans


def _read_plan_lines(wb, plans, source, result) -> list[dict]:
    if "6_MaintPlanLines" not in wb.sheetnames:
        result["errors"].append("Missing sheet 6_MaintPlanLines")
        return []
    rows = list(wb["6_MaintPlanLines"].iter_rows(values_only=True))
    hdr_idx, header = _find_header_row(rows, "Plan ID", "Interval unit")
    if header is None:
        result["errors"].append("6_MaintPlanLines: header row not found")
        return []
    ci = {
        "plan": _header_index(header, "Plan ID"),
        "asset_or_fl": _header_index(header, "Asset or Functional location"),
        "ref_type": _header_index(header, "Reference type"),
        "ref_value": _header_index(header, "Reference value"),
        "interval_unit": _header_index(header, "Interval unit"),
        "interval_value": _header_index(header, "Interval value"),
        "start_date": _header_index(header, "Start date"),
        "job_type": _header_index(header, "Job type"),
        "job_desc": _header_index(header, "Job description"),
    }
    lines = []
    for row in rows[hdr_idx + 1:]:
        def cell(key):
            i = ci[key]
            if i is None or i >= len(row):
                return ""
            return row[i] if key == "start_date" else _txt(row[i])
        plan_id = cell("plan")
        if not plan_id:
            continue
        plan = plans.get(plan_id, {})
        lines.append({
            "source": source,
            "planId": plan_id,
            "planName": plan.get("name", ""),
            "workOrderType": plan.get("workOrderType", ""),
            "trade": plan.get("trade", ""),
            "active": plan.get("active", True),
            "assetOrFL": cell("asset_or_fl"),
            "refType": cell("ref_type"),
            "refValue": cell("ref_value"),
            "intervalUnit": cell("interval_unit"),
            "intervalValue": cell("interval_value"),
            "startDate": _coerce_date(cell("start_date")),
            "jobType": cell("job_type"),
            "jobDescription": cell("job_desc"),
        })
    return lines


# ── master / classification sources ──────────────────────────────────────────
def read_master(path: str | Path) -> dict:
    """Read the classification sources from Asset_Master.xlsx.

    Returns {"assets": {ID: {...}}, "feedMap": {CODE: {...}}, "scopeMap":
    {group: scope}, "errors": [...]}. Missing sheets/columns are reported in
    errors rather than raised, so the verifier can list exactly what the master
    still needs.
    """
    path = Path(path)
    out = {"assets": {}, "feedMap": {}, "scopeMap": {}, "errors": []}
    if not path.exists():
        out["errors"].append(f"Asset_Master not found: {path}")
        return out
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        out["assets"] = _read_asset_master_sheet(wb, out)
        out["feedMap"] = _read_pm_feed_map_sheet(wb, out)
        out["scopeMap"] = _read_scope_map(wb, out)
    finally:
        wb.close()
    return out


def _read_asset_master_sheet(wb, out) -> dict:
    if "Asset_Master" not in wb.sheetnames:
        out["errors"].append("Asset_Master: missing 'Asset_Master' sheet")
        return {}
    rows = list(wb["Asset_Master"].iter_rows(values_only=True))
    hdr_idx, header = _find_header_row(rows, "Asset ID", "Stage")
    if header is None:
        out["errors"].append("Asset_Master: header row not found")
        return {}
    ci = {
        "id": _header_index(header, "Asset ID"),
        "name": _header_index(header, "Asset Name"),
        "stage": _header_index(header, "Stage"),
        "main": _header_index(header, "Category", "Main Asset Group"),
        "sub": _header_index(header, "Sub Asset Group"),
        "loc": _header_index(header, "Location"),
        "sys": _header_index(header, "System/Area", "System / Area"),
        "pic": _header_index(header, "PIC / Owner", "PIC/Owner", "PIC", "Default PIC"),
    }
    if ci["pic"] is None:
        out["errors"].append("Asset_Master: no 'PIC / Owner' column (PIC will be blank for asset-resolved lines)")
    assets = {}
    for row in rows[hdr_idx + 1:]:
        def cell(key):
            i = ci[key]
            return _txt(row[i]) if i is not None and i < len(row) else ""
        aid = cell("id")
        if not aid:
            continue
        assets[aid.upper()] = {
            "assetName": cell("name"),
            "stage": cell("stage"),
            "mainGroup": cell("main"),
            "subGroup": cell("sub"),
            "location": cell("loc"),
            "systemArea": cell("sys"),
            "pic": cell("pic"),
        }
    return assets


def _read_pm_feed_map_sheet(wb, out) -> dict:
    sheet = next((s for s in wb.sheetnames if _norm_header(s) in ("pmfeedmap", "feedmap")), None)
    if sheet is None:
        out["errors"].append("Asset_Master: missing 'PM_Feed_Map' sheet (functional-location codes cannot be classified)")
        return {}
    rows = list(wb[sheet].iter_rows(values_only=True))
    hdr_idx, header = _find_header_row(rows, "Feed Code", "Stage")
    if header is None:
        out["errors"].append("PM_Feed_Map: header row not found")
        return {}
    ci = {
        "code": _header_index(header, "Feed Code"),
        "type": _header_index(header, "Code Type"),
        "stage": _header_index(header, "Stage"),
        "scope": _header_index(header, "Scope"),
        "sys": _header_index(header, "System / Area", "System/Area"),
        "loc": _header_index(header, "Location (display)", "Location"),
        "pic": _header_index(header, "Default PIC", "PIC"),
    }
    feed_map = {}
    for row in rows[hdr_idx + 1:]:
        def cell(key):
            i = ci[key]
            return _txt(row[i]) if i is not None and i < len(row) else ""
        code = cell("code")
        if not code:
            continue
        feed_map[code.upper()] = {
            "codeType": cell("type"),
            "stage": cell("stage"),
            "scope": cell("scope"),
            "systemArea": cell("sys"),
            "location": cell("loc"),
            "pic": cell("pic"),
        }
    return feed_map


def _read_scope_map(wb, out) -> dict:
    """Lists sheet: a 'Category (for Scope map)' column -> 'Scope' column.
    Also accepts the legacy column name 'Main Asset Group' for backward compatibility."""
    if "Lists" not in wb.sheetnames:
        out["errors"].append("Asset_Master: missing 'Lists' sheet (no scope map)")
        return {}
    rows = list(wb["Lists"].iter_rows(values_only=True))
    hdr_idx, header = _find_header_row(rows, "Category", "Scope")
    if header is None:
        hdr_idx, header = _find_header_row(rows, "Main Asset Group", "Scope")
    if header is None:
        out["errors"].append("Lists: no Category -> Scope map (need the 'Category' + 'Scope' columns)")
        return {}
    gi = _header_index(header, "Category", "Main Asset Group")
    si = _header_index(header, "Scope")
    scope_map = {}
    for row in rows[hdr_idx + 1:]:
        group = _txt(row[gi]) if gi is not None and gi < len(row) else ""
        scope = _txt(row[si]) if si is not None and si < len(row) else ""
        if group and scope:
            scope_map[group.strip().lower()] = scope
    return scope_map


# ── resolution (most-specific-first) ─────────────────────────────────────────
def _names_asset_id(line: dict) -> str:
    """Return the Asset ID this line names, or "" if it is a feed-code line."""
    a = line.get("assetOrFL", "")
    if ASSET_ID_RE.match(a):
        return a
    if line.get("refType", "").strip().lower() == "asset":
        return line.get("refValue", "")
    return ""


def resolve_line(line: dict, master: dict) -> dict:
    """Resolve one plan line to {stage, scope, systemArea, location, pic,
    assetCode, assetName, mainGroup, mappingPath, unresolved}."""
    assets = master.get("assets", {})
    feed_map = master.get("feedMap", {})
    scope_map = master.get("scopeMap", {})

    res = {"stage": "", "scope": "", "systemArea": "", "location": "", "pic": "",
           "assetCode": "", "assetName": "", "mainGroup": "", "mappingPath": "unresolved"}

    asset_id = _names_asset_id(line)
    a = assets.get(asset_id.upper()) if asset_id else None
    feed_code = line.get("refValue", "")
    f = feed_map.get(feed_code.upper()) if feed_code else None

    # (a) Asset_Master is most specific.
    if a:
        res["assetCode"] = asset_id
        res["assetName"] = a.get("assetName", "")
        res["stage"] = a.get("stage", "")
        res["mainGroup"] = a.get("mainGroup", "")
        res["systemArea"] = a.get("systemArea", "")
        res["location"] = a.get("location", "")
        res["pic"] = a.get("pic", "")
        if res["mainGroup"]:
            res["scope"] = scope_map.get(res["mainGroup"].strip().lower(), "")
        res["mappingPath"] = "asset_master"

    # (b) PM_Feed_Map (primary for feed-code lines; fill for asset lines).
    if f:
        if res["mappingPath"] == "unresolved":
            res["mappingPath"] = "pm_feed_map"
            res["assetCode"] = asset_id or feed_code
        res["stage"] = res["stage"] or f.get("stage", "")
        res["scope"] = res["scope"] or f.get("scope", "")
        res["systemArea"] = res["systemArea"] or f.get("systemArea", "")
        res["location"] = res["location"] or f.get("location", "")
        res["pic"] = res["pic"] or f.get("pic", "")

    # (c) last-resort identity fill so cards never show empty asset code.
    if not res["assetCode"]:
        res["assetCode"] = asset_id or feed_code
    if not res["assetName"]:
        res["assetName"] = line.get("planName", "") or res["assetCode"]
    return res


# ── task building ────────────────────────────────────────────────────────────
def _status_for(scheduled: date, today: date) -> str:
    return "overdue" if scheduled < today else "pending"


def build_pm_tasks(feeds: list[dict], master: dict, opts: dict | None = None) -> list[dict]:
    """Port of buildPMTasks(feeds, master, opts).

    feeds: list of {"path": ..., "source": "production"|"utility"}.
    master: dict from read_master().
    opts: {"year": int, "today": date, "win_start": date, "win_end": date}.
    """
    opts = opts or {}
    today = opts.get("today") or date.today()
    year = int(opts.get("year") or today.year)
    win_start = opts.get("win_start") or date(year, 1, 1)
    win_end = opts.get("win_end") or date(year, 12, 31)

    tasks: list[dict] = []
    for feed in feeds:
        parsed = read_feed(feed["path"], feed["source"])
        for line in parsed["lines"]:
            if not line["active"]:
                continue
            res = resolve_line(line, master)
            occurrences = expand_occurrences(
                line["startDate"], line["intervalUnit"], line["intervalValue"], win_start, win_end,
            )
            for occ in occurrences:
                scheduled = occ.isoformat()
                tasks.append({
                    "id": f"feed-{line['source']}-{line['planId']}-{scheduled}",
                    "planId": line["planId"],
                    "source": f"D365 {line['source'].capitalize()} Feed",
                    "assetCode": res["assetCode"],
                    "assetName": res["assetName"],
                    "stage": res["stage"],
                    "scope": res["scope"],
                    "systemArea": res["systemArea"],
                    "location": res["location"],
                    "pic": res["pic"],
                    "scheduled": scheduled,
                    "completion": None,              # MANUAL only (Mark Done)
                    "status": _status_for(occ, today),
                    "intervalUnit": line["intervalUnit"],
                    "intervalValue": line["intervalValue"],
                    "jobType": line["jobType"],
                    "jobDescription": line["jobDescription"],
                    # diagnostics (not for display)
                    "_mappingPath": res["mappingPath"],
                    "_mainGroup": res["mainGroup"],
                })
    return tasks


# ── adapter to the dashboard's internal task shape ───────────────────────────
_MONTH_LABELS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                 "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _frequency_label(unit: str, value) -> str:
    try:
        v = int(value)
    except (TypeError, ValueError):
        return ""
    u = (unit or "").strip().lower()
    if u.startswith("month"):
        return {1: "Monthly", 2: "Bi-Monthly", 3: "Quarterly", 4: "4-Monthly",
                6: "Half-Yearly", 12: "Yearly"}.get(v, f"Every {v} Months")
    if u.startswith("year"):
        return "Yearly" if v == 1 else f"Every {v} Years"
    if u.startswith("week"):
        return "Weekly" if v == 1 else f"Every {v} Weeks"
    if u.startswith("day"):
        return "Daily" if v == 1 else f"Every {v} Days"
    return f"Every {v} {unit}".strip()


def _internal_scope(scope_display: str) -> str:
    """Map the Lists Scope label to the dashboard's internal scope key."""
    return "equipment" if (scope_display or "").strip().lower() == "production equipment" else "utility"


def build_feed_tasks_internal(feeds: list[dict], master: dict, opts: dict | None = None) -> list[dict]:
    """Emit feed tasks in the dashboard's INTERNAL task shape so they merge into
    pm_schedule_service's task list. The override layer derives the operational
    booleans (isDone/isOverdue/…) from status + plannedDate, so completion stays
    manual-only here."""
    opts = opts or {}
    today = opts.get("today") or date.today()
    year = int(opts.get("year") or today.year)
    win_start = opts.get("win_start") or date(year, 1, 1)
    win_end = opts.get("win_end") or date(year, 12, 31)
    # Stage is driven by the master (Asset_Master / PM_Feed_Map). stage_override is
    # an optional hard override (normally unset); DEFAULT_FEED_STAGE is the blank
    # fallback. Scope/System-Area/Location/PIC always come from the workbook.
    stage_override = opts.get("stage_override")

    out: list[dict] = []
    for feed in feeds:
        parsed = read_feed(feed["path"], feed["source"])
        is_prod = feed["source"] == PRODUCTION_FEED
        src_label = "D365 Production Feed" if is_prod else "D365 Utility Feed"
        src_slot = "feed_production" if is_prod else "feed_utility"
        src_file = Path(feed["path"]).name
        for line in parsed["lines"]:
            if not line["active"]:
                continue
            res = resolve_line(line, master)
            scope_display = res["scope"]
            scope_internal = _internal_scope(scope_display)
            main_group = res["mainGroup"] or scope_display or "Utilities"
            freq = _frequency_label(line["intervalUnit"], line["intervalValue"])
            desc = line["jobDescription"] or line["planName"] or (f"{freq} Preventive Maintenance".strip())
            mapping_status = {
                "asset_master": "Mapped",
                "pm_feed_map": "Feed-mapped",
            }.get(res["mappingPath"], "Unmapped")
            for occ in expand_occurrences(line["startDate"], line["intervalUnit"], line["intervalValue"], win_start, win_end):
                iso = occ.isoformat()
                out.append({
                    "pmTaskId": f"feed-{line['source']}-{line['planId']}-{iso}",
                    "planId": line["planId"],
                    "stage": stage_override or res["stage"] or DEFAULT_FEED_STAGE,
                    "assetId": res["assetCode"],
                    "assetName": res["assetName"],
                    "mainAssetGroup": main_group,
                    "subAssetGroup": "",
                    "systemArea": res["systemArea"] or "Unassigned",
                    "location": res["location"] or "Unassigned",
                    "pmDescription": desc,
                    "frequency": freq,
                    "plannedYear": occ.year,
                    "plannedMonth": occ.month,
                    "plannedMonthLabel": _MONTH_LABELS[occ.month - 1],
                    "plannedQuarter": f"Q{(occ.month - 1) // 3 + 1}",
                    "plannedDate": iso,
                    "plannedDateLabel": occ.strftime("%d %b %Y"),
                    "contractorOrPIC": res["pic"],
                    "provider": "",
                    "scheduleStatus": "Overdue" if occ < today else "Not Due",
                    "completionStatus": "Open",
                    "actualCompletionDate": None,
                    "daysOverdue": 0,
                    "sourceFile": src_file,
                    "sourceSlot": src_slot,
                    "sourceLabel": src_label,
                    "sourceSheet": "6_MaintPlanLines",
                    "mappingStatus": mapping_status,
                    "domain": "equipment" if scope_internal == "equipment" else "utility",
                    "scope": scope_internal,
                    "scopeLabel": scope_display,
                    "jobType": line["jobType"],
                    "intervalUnit": line["intervalUnit"],
                    "intervalValue": line["intervalValue"],
                    "isDone": False,
                    "isDueThisMonth": False,
                    "isDueSoon": False,
                    "isOverdue": occ < today,
                    "needsReview": False,
                })
    return out


# ── verification harness ─────────────────────────────────────────────────────
def verify(feeds: list[dict], master_path: str | Path, opts: dict | None = None) -> dict:
    """Build tasks and report any classification gaps (used before wiring)."""
    master = read_master(master_path)
    tasks = build_pm_tasks(feeds, master, opts)
    blank_stage = [t for t in tasks if not t["stage"]]
    blank_scope = [t for t in tasks if not t["scope"]]
    blank_sys = [t for t in tasks if not t["systemArea"]]
    unresolved = sorted({
        (t["planId"], t["assetCode"]) for t in tasks if t["_mappingPath"] == "unresolved"
    })
    return {
        "masterErrors": master["errors"],
        "taskCount": len(tasks),
        "blankStage": len(blank_stage),
        "blankScope": len(blank_scope),
        "blankSystemArea": len(blank_sys),
        "unresolvedLines": unresolved,
        "tasks": tasks,
    }
