"""
PM schedule source registry, local source helpers, and Stage 2 utility parsing.

This module keeps the PM schedule tracker flexible by defining explicit local
source slots for:
- Stage 1 Utility (active)
- Stage 1 Production Equipment (active)
- Stage 2 Utility (active)
- Stage 2 Production Equipment (active)
"""

from __future__ import annotations

import copy
import json
import re
import shutil
from collections import OrderedDict
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

from maintenance_service import (
    DATA_DIR,
    EQUIPMENT_MAINTENANCE_COPY_PATH,
    UTILITY_MAINTENANCE_SOURCE_PATH,
    clean_text,
    clear_maintenance_caches,
    contains_thai,
    find_equipment_header_row,
    get_file_signature,
    get_sheet_due_dates,
    is_schedule_marker,
    normalize_asset_code,
    parse_month_token,
    parse_week_token,
    resolve_equipment_stage1_workbook_path,
    resolve_utility_workbook_path,
)
from downtime_service import translate_maintenance_description

PM_SCHEDULE_UPLOAD_EXTENSIONS = {".xlsx", ".xls"}
PM_SCHEDULE_SOURCE_REGISTRY_PATH = DATA_DIR / "pm_schedule_source_registry.json"
PM_SCHEDULE_SOURCE_ARCHIVE_DIR = DATA_DIR / "pm_schedule_imports"

PM_SCHEDULE_SOURCE_SPECS = OrderedDict(
    [
        (
            "utility_stage1",
            {
                "label": "Stage 1 Utility",
                "scope": "utility",
                "default_stage": "Stage 1",
                "template": "legacy_utility",
                "template_label": "Legacy utility UL schedule",
                "template_hint": "Workbook sheets with 'Machine Code' and 'Machine Name' headers.",
                "description": "Active Stage 1 utility PM schedule used by the legacy utility tracking views.",
                "canonical_path": UTILITY_MAINTENANCE_SOURCE_PATH,
                "resolver": resolve_utility_workbook_path,
                "default_active": True,
                "supports_manual_activation": False,
            },
        ),
        (
            "equipment_stage1",
            {
                "label": "Stage 1 Production Equipment",
                "scope": "equipment",
                "default_stage": "Stage 1",
                "template": "modern_equipment",
                "template_label": "Production equipment PM layout",
                "template_hint": "Workbook sheets with 'Machine Code' / 'Machine Name' headers and week-number schedule columns.",
                "description": "Confirmed Stage 1 production equipment PM schedule workbook.",
                "canonical_path": EQUIPMENT_MAINTENANCE_COPY_PATH,
                "resolver": resolve_equipment_stage1_workbook_path,
                "default_active": True,
                "supports_manual_activation": False,
            },
        ),
        # Stage 2 PM is now produced from the live D365 PM feed
        # (see pm_feed_integration), not from a registry source slot. The old
        # Stage 2 utility/equipment slots and their generators were removed.
    ]
)

_UTILITY_GROUP_KEYWORDS = (
    ("refriger", "Refrigeration"),
    ("compressor", "Air Compressor"),
    ("hot water", "Hot Water"),
    ("fire", "Fire Protection"),
    ("water treatment", "Water Treatment"),
    ("raw water", "Water Supply"),
    ("boiler", "Boiler"),
    ("wastewater", "Wastewater"),
    ("waste water", "Wastewater"),
    ("lpg", "LPG"),
    ("foam", "Foam Cleaning"),
    ("lift", "Lift"),
)
_SOURCE_STATUS_CACHE = {"signature": None, "value": None}


def _spec(slot_key: str) -> dict:
    spec = PM_SCHEDULE_SOURCE_SPECS.get(slot_key)
    if not spec:
        raise ValueError(f"Unknown PM schedule source slot: {slot_key}")
    return spec


def _default_registry() -> dict:
    return {
        "sources": {
            key: {
                "active": bool(spec.get("default_active")),
                "file_name": None,
                "last_uploaded_at": None,
                "validation": {},
            }
            for key, spec in PM_SCHEDULE_SOURCE_SPECS.items()
        }
    }


def _load_registry() -> dict:
    payload = _default_registry()
    try:
        if PM_SCHEDULE_SOURCE_REGISTRY_PATH.exists():
            stored = json.loads(PM_SCHEDULE_SOURCE_REGISTRY_PATH.read_text(encoding="utf-8"))
            if isinstance(stored, dict):
                stored_sources = stored.get("sources") or {}
                for key in PM_SCHEDULE_SOURCE_SPECS:
                    if isinstance(stored_sources.get(key), dict):
                        payload["sources"][key].update(stored_sources[key])
    except Exception:
        pass
    return payload


def _save_registry(payload: dict):
    PM_SCHEDULE_SOURCE_REGISTRY_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _registry_entry(slot_key: str, registry: dict | None = None) -> dict:
    current = registry or _load_registry()
    return current.setdefault("sources", {}).setdefault(
        slot_key,
        {
            "active": bool(_spec(slot_key).get("default_active")),
            "file_name": None,
            "last_uploaded_at": None,
            "validation": {},
        },
    )


def _resolve_source_path(slot_key: str) -> Path | None:
    spec = _spec(slot_key)
    canonical_path = spec["canonical_path"]
    if canonical_path.exists():
        return canonical_path
    resolver = spec.get("resolver")
    if resolver is None:
        return None
    try:
        candidate = resolver()
    except Exception:
        return None
    candidate_path = Path(candidate) if candidate else None
    if candidate_path and candidate_path.exists():
        return candidate_path
    return None


def _same_path(left: Path | None, right: Path | None) -> bool:
    if not left or not right:
        return False
    try:
        return left.resolve() == right.resolve()
    except OSError:
        return left == right


def _display_path(path: Path | None) -> str | None:
    if not path:
        return None
    try:
        return path.relative_to(DATA_DIR.parent).as_posix()
    except ValueError:
        return str(path)


def _parse_iso_timestamp(value) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value))
    except (TypeError, ValueError):
        return None


def _latest_timestamp(values) -> str | None:
    parsed = [item for item in (_parse_iso_timestamp(value) for value in values) if item is not None]
    if not parsed:
        return None
    return max(parsed).isoformat()


def _find_legacy_utility_sheet_count(workbook) -> int:
    matches = 0
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        max_row = min(8, worksheet.max_row)
        max_col = min(6, worksheet.max_column)
        for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True):
            values = {
                clean_text(value)
                for value in row
                if clean_text(value)
            }
            if "Machine Code" in values and "Machine Name" in values:
                matches += 1
                break
    return matches


def _extract_modern_utility_columns(worksheet) -> list[dict]:
    if clean_text(worksheet.cell(3, 1).value) != "ITEM":
        return []
    if clean_text(worksheet.cell(3, 2).value) != "DESCRIPTION":
        return []

    columns = []
    current_month = None
    for column_index in range(7, worksheet.max_column + 1):
        month_value = parse_month_token(worksheet.cell(4, column_index).value) or parse_month_token(worksheet.cell(3, column_index).value)
        if month_value:
            current_month = month_value
        week_value = parse_week_token(worksheet.cell(5, column_index).value)
        if current_month and week_value:
            columns.append(
                {
                    "column_index": column_index,
                    "month": current_month,
                    "target_week": week_value,
                }
            )
    return columns


def _load_modern_utility_group_lookup(workbook) -> dict[str, dict]:
    lookup = {}
    for sheet_name in workbook.sheetnames:
        lowered = str(sheet_name).strip().lower()
        if "group" not in lowered:
            continue
        worksheet = workbook[sheet_name]
        current_group = ""
        current_location = ""
        for row in worksheet.iter_rows(values_only=True):
            asset_id_raw = clean_text(row[0] if len(row) > 0 else None)
            machine_code = normalize_asset_code(row[1] if len(row) > 1 else None)
            asset_name = clean_text(row[2] if len(row) > 2 else None)
            location_hint = clean_text(row[3] if len(row) > 3 else None)

            if asset_id_raw and asset_id_raw.upper() not in {"ASSET*", "UT"} and asset_id_raw.rstrip().endswith(":"):
                current_group = asset_id_raw.rstrip(":").strip()
                current_location = asset_name or location_hint or current_location
                continue

            if machine_code:
                lookup[machine_code] = {
                    "asset_id": normalize_asset_code(asset_id_raw) or machine_code,
                    "asset_name": asset_name,
                    "group": current_group,
                    "location": current_location or location_hint or "",
                }
    return lookup


def _count_modern_utility_marked_rows(worksheet, schedule_columns: list[dict]) -> int:
    marked_rows = 0
    for row in worksheet.iter_rows(min_row=6, values_only=True):
        code_text = clean_text(row[1] if len(row) > 1 else None)
        name_text = clean_text(row[2] if len(row) > 2 else None)
        if not code_text and not name_text:
            continue
        if any(
            len(row) >= item["column_index"] and is_schedule_marker(row[item["column_index"] - 1])
            for item in schedule_columns
        ):
            marked_rows += 1
    return marked_rows


def _validate_legacy_utility_workbook(path: Path) -> dict:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    schedule_sheet_count = _find_legacy_utility_sheet_count(workbook)
    if not schedule_sheet_count:
        raise ValueError("No legacy utility schedule sheets with 'Machine Code' and 'Machine Name' headers were found.")
    return {
        "sheet_count": len(workbook.sheetnames),
        "schedule_sheet_count": schedule_sheet_count,
        "template": "legacy_utility",
    }


def _validate_modern_utility_workbook(path: Path) -> dict:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    schedule_sheet_names = []
    marked_row_count = 0
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        schedule_columns = _extract_modern_utility_columns(worksheet)
        if not schedule_columns:
            continue
        schedule_sheet_names.append(sheet_name)
        marked_row_count += _count_modern_utility_marked_rows(worksheet, schedule_columns)
    if not schedule_sheet_names:
        raise ValueError("No local utility schedule sheets with the expected Schedule (PM) layout were found.")
    return {
        "sheet_count": len(workbook.sheetnames),
        "schedule_sheet_count": len(schedule_sheet_names),
        "schedule_sheets": schedule_sheet_names[:6],
        "group_lookup_count": len(_load_modern_utility_group_lookup(workbook)),
        "marked_row_count": marked_row_count,
        "template": "modern_utility",
    }


def _validate_modern_equipment_workbook(path: Path) -> dict:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    header_sheet_count = sum(1 for sheet_name in workbook.sheetnames if find_equipment_header_row(workbook[sheet_name]))
    if not header_sheet_count:
        raise ValueError("No local equipment schedule sheets with 'Machine Code' / 'Machine Name' headers were found.")
    return {
        "sheet_count": len(workbook.sheetnames),
        "header_sheet_count": header_sheet_count,
        "template": "modern_equipment",
    }


def _validate_source_path(slot_key: str, path: Path) -> dict:
    template = _spec(slot_key).get("template")
    if template == "legacy_utility":
        return _validate_legacy_utility_workbook(path)
    if template == "modern_utility":
        return _validate_modern_utility_workbook(path)
    if template == "modern_equipment":
        return _validate_modern_equipment_workbook(path)
    raise ValueError(f"Unsupported PM schedule template type: {template}")


def _build_source_message(
    *,
    available: bool,
    active: bool,
    supports_manual_activation: bool,
    slot_key: str,
    editable_label: str | None = None,
    validation_error: str | None = None,
) -> str:
    if validation_error:
        return validation_error
    if not available:
        if editable_label:
            return f"Source file is not available yet. Add or update the workbook at {editable_label}."
        return "Source file is not available."
    if active:
        return "Included in PM tracking."
    if supports_manual_activation:
        return "Workbook detected and currently staged outside KPI tracking."
    return "Available."


def get_pm_schedule_source_status() -> OrderedDict:
    source_paths = OrderedDict((slot_key, _resolve_source_path(slot_key)) for slot_key in PM_SCHEDULE_SOURCE_SPECS)
    cache_signature = (
        get_file_signature(PM_SCHEDULE_SOURCE_REGISTRY_PATH),
        tuple((slot_key, get_file_signature(path) if path else None) for slot_key, path in source_paths.items()),
    )
    if _SOURCE_STATUS_CACHE["signature"] == cache_signature and _SOURCE_STATUS_CACHE["value"] is not None:
        return copy.deepcopy(_SOURCE_STATUS_CACHE["value"])

    registry = _load_registry()
    status = OrderedDict()
    for slot_key, spec in PM_SCHEDULE_SOURCE_SPECS.items():
        entry = _registry_entry(slot_key, registry)
        path = source_paths.get(slot_key)
        editable_path = spec["canonical_path"]
        available = bool(path and path.exists())
        active = bool(entry.get("active")) if spec.get("supports_manual_activation") else available
        validation = {}
        validation_error = None
        if available:
            try:
                validation = _validate_source_path(slot_key, path)
            except Exception as exc:
                validation_error = str(exc)
        status[slot_key] = {
            "slot": slot_key,
            "label": spec["label"],
            "scope": spec["scope"],
            "default_stage": spec["default_stage"],
            "template": spec["template"],
            "template_label": spec["template_label"],
            "template_hint": spec["template_hint"],
            "description": spec["description"],
            "editable_path": str(editable_path),
            "editable_path_label": _display_path(editable_path),
            "available": available,
            "active": bool(active and available),
            "supports_manual_activation": bool(spec.get("supports_manual_activation")),
            "file_name": path.name if available else entry.get("file_name"),
            "path": str(path) if available else None,
            "path_label": _display_path(path) if available else None,
            "path_mode": "editable" if _same_path(path, editable_path) else ("fallback" if available else "missing"),
            "last_modified": datetime.fromtimestamp(path.stat().st_mtime).isoformat() if available else None,
            "last_uploaded_at": entry.get("last_uploaded_at"),
            "validation": validation or entry.get("validation") or {},
            "message": _build_source_message(
                available=available,
                active=bool(active and available),
                supports_manual_activation=bool(spec.get("supports_manual_activation")),
                slot_key=slot_key,
                editable_label=_display_path(editable_path),
                validation_error=validation_error,
            ),
        }
    _SOURCE_STATUS_CACHE["signature"] = cache_signature
    _SOURCE_STATUS_CACHE["value"] = copy.deepcopy(status)
    return status


def summarize_pm_schedule_sources(source_map, tracked_counts: dict | None = None) -> dict:
    tracked_counts = tracked_counts or {}
    latest_source_update = _latest_timestamp(source.get("last_modified") for source in source_map.values())
    for slot_key, source in source_map.items():
        source["tracked_task_count"] = int(tracked_counts.get(slot_key, 0))
    return {
        "slotCount": len(source_map),
        "availableCount": sum(1 for source in source_map.values() if source.get("available")),
        "activeCount": sum(1 for source in source_map.values() if source.get("active")),
        "stagedCount": sum(1 for source in source_map.values() if source.get("available") and not source.get("active")),
        "missingCount": sum(1 for source in source_map.values() if not source.get("available")),
        "trackedTaskCount": sum(int(value or 0) for value in tracked_counts.values()),
        "latestSourceUpdate": latest_source_update,
    }


def get_pm_schedule_last_synced() -> str | None:
    return summarize_pm_schedule_sources(get_pm_schedule_source_status()).get("latestSourceUpdate")


def _remove_existing_variants(path: Path):
    for suffix in (".xlsx", ".xls"):
        candidate = path.with_suffix(suffix)
        if candidate.exists():
            try:
                candidate.unlink()
            except OSError:
                pass


def _stage_uploaded_file(file_storage, fallback_stem: str) -> Path:
    filename = Path(getattr(file_storage, "filename", "") or "").name
    extension = Path(filename).suffix.lower()
    if extension not in PM_SCHEDULE_UPLOAD_EXTENSIONS:
        raise ValueError("Unsupported file type. Upload an XLSX or XLS maintenance schedule.")
    upload_dir = DATA_DIR / "_upload_tmp"
    upload_dir.mkdir(parents=True, exist_ok=True)
    safe_stem = re.sub(r"[^A-Za-z0-9._-]+", "_", Path(filename).stem).strip("._-") or fallback_stem
    temp_path = upload_dir / f"{safe_stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{extension}"
    file_storage.save(temp_path)
    return temp_path


def _promote_uploaded_schedule_file(temp_path: Path, canonical_path: Path, archive_dir: Path) -> Path:
    _remove_existing_variants(canonical_path)
    final_path = canonical_path.with_suffix(temp_path.suffix.lower())
    final_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path.replace(final_path)
    archive_dir.mkdir(parents=True, exist_ok=True)
    archive_path = archive_dir / f"{final_path.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{final_path.suffix.lower()}"
    shutil.copy2(final_path, archive_path)
    return final_path


def import_pm_schedule_source_file(file_storage, slot_key: str, activate: bool | None = None) -> dict:
    temp_path = None
    try:
        spec = _spec(slot_key)
        temp_path = _stage_uploaded_file(file_storage, spec["canonical_path"].stem)
        validation = _validate_source_path(slot_key, temp_path)
        final_path = _promote_uploaded_schedule_file(
            temp_path,
            spec["canonical_path"],
            PM_SCHEDULE_SOURCE_ARCHIVE_DIR / slot_key,
        )

        registry = _load_registry()
        entry = _registry_entry(slot_key, registry)
        if spec.get("supports_manual_activation"):
            if activate is None:
                next_active = bool(entry.get("active"))
            else:
                next_active = bool(activate)
        else:
            next_active = True
        entry.update(
            {
                "active": next_active,
                "file_name": final_path.name,
                "last_uploaded_at": datetime.now().isoformat(),
                "validation": validation,
            }
        )
        _save_registry(registry)
        clear_maintenance_caches()
        _SOURCE_STATUS_CACHE["signature"] = None
        _SOURCE_STATUS_CACHE["value"] = None
        return {
            "ok": True,
            "message": f"{spec['label']} schedule uploaded.",
            "slot": slot_key,
            "file": final_path.name,
            "active": next_active,
            "validation_summary": validation,
        }
    except Exception as exc:
        if temp_path and temp_path.exists():
            try:
                temp_path.unlink()
            except OSError:
                pass
        return {"ok": False, "message": f"PM schedule upload failed: {exc}"}


def set_pm_schedule_source_active(slot_key: str, active: bool) -> dict:
    try:
        spec = _spec(slot_key)
        if not spec.get("supports_manual_activation"):
            return {"ok": False, "message": f"{spec['label']} is always active when available."}

        path = _resolve_source_path(slot_key)
        if not path:
            return {"ok": False, "message": f"{spec['label']} has no uploaded workbook to activate yet."}

        registry = _load_registry()
        entry = _registry_entry(slot_key, registry)
        entry["active"] = bool(active)
        _save_registry(registry)
        clear_maintenance_caches()
        _SOURCE_STATUS_CACHE["signature"] = None
        _SOURCE_STATUS_CACHE["value"] = None
        return {
            "ok": True,
            "message": f"{spec['label']} {'activated' if active else 'kept staged'} for PM tracking.",
            "slot": slot_key,
            "active": bool(active),
        }
    except Exception as exc:
        return {"ok": False, "message": f"PM schedule source activation failed: {exc}"}


def _looks_like_schedule_code(text: str | None) -> bool:
    value = normalize_asset_code(text)
    if not value:
        return False
    return bool(re.fullmatch(r"[A-Z0-9][A-Z0-9/_-]*", value))


def _safe_schedule_identifier(prefix: str, label: str, row_index: int) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", (label or "").lower()).strip("-")
    slug = slug[:40] or f"row-{row_index}"
    return f"{prefix}-{slug}".upper()


def _frequency_label_from_times_per_year(value) -> str:
    text = clean_text(value)
    if not text:
        return "Scheduled"
    try:
        count = int(float(text))
    except (TypeError, ValueError):
        return text
    if count == 12:
        return "Monthly"
    if count == 6:
        return "Bi-monthly"
    if count == 4:
        return "Quarterly"
    if count == 2:
        return "Semi-annual"
    if count == 1:
        return "Annual"
    return f"{count}x / year"


# NOTE: the Stage 2 utility generator (build_stage2_utility_dataset and its
# _extract_modern_utility_occurrences fabricator) was removed. Stage 2 PM now
# comes from the live D365 PM feed (see pm_feed_integration).
