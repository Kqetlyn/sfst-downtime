import json
import logging
import os
import re
import threading
from glob import glob
from datetime import datetime, timedelta

import pandas as pd
import openpyxl

try:
    from deep_translator import GoogleTranslator as _GoogleTranslator
    _DEEP_TRANSLATOR_AVAILABLE = True
except ImportError:
    _DEEP_TRANSLATOR_AVAILABLE = False

from downtime_management import (
    build_management_downtime_payload,
    enrich_work_order_records,
    CRITICALITY_CRITICAL,
    CRITICALITY_NON_CRITICAL,
    CRITICALITY_RANK,
    REFRIGERATION_GROUP,
    _normalize_criticality,
    _normalize_display_criticality,
)
from asset_mapping import get_asset_mapping_meta as get_grouped_machine_mapping_meta

_log = logging.getLogger(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.abspath(os.path.join(BASE_DIR, "..", "data"))
# Standalone: cache written to data/ since there is no Downtime frontend folder here
DOWNTIME_CACHE_OUTPUT_FILE = os.path.abspath(
    os.path.join(BASE_DIR, "..", "data", "downtime-cache.json")
)
ARGOS_TRANSLATE_DIR = os.path.join(DATA_DIR, "argos_translate")
os.environ.setdefault("XDG_CONFIG_HOME", os.path.join(ARGOS_TRANSLATE_DIR, "config"))
os.environ.setdefault("XDG_DATA_HOME", os.path.join(ARGOS_TRANSLATE_DIR, "data"))
os.environ.setdefault("XDG_CACHE_HOME", os.path.join(ARGOS_TRANSLATE_DIR, "cache"))
os.environ.setdefault("ARGOS_STANZA_AVAILABLE", "0")
os.environ.setdefault("ARGOS_CHUNK_TYPE", "MINISBD")

_DOWNTIME_CACHE = {}
_WO_LOAD_CACHE = {"sig": None, "payload": None}
# Keyed by normalised stage string; cleared by import_work_order_file().
_SQL_WO_CACHE: dict = {}
DOWNTIME_CACHE_VERSION = "2026-06-18-stage-text-detection"
DOWNTIME_EXPORT_YEAR = 2026

PRIMARY_WORK_ORDER_DOWNTIME_FILE = os.path.join(DATA_DIR, "data downtime.csv")
FALLBACK_WORK_ORDER_DOWNTIME_FILE = os.path.abspath(os.path.join(os.path.expanduser("~"), "Downloads", "data downtime.csv"))
WORK_ORDER_IMPORT_DIR = os.path.join(DATA_DIR, "work_order_imports")
WORK_ORDER_IMPORT_EXTENSIONS = {".csv", ".xlsx", ".xls"}
ASSET_MASTER_FILENAME = "Asset_Master.xlsx"
ASSET_MASTER_RELATIVE_PATH = os.path.join("master", ASSET_MASTER_FILENAME)
SLA_TARGETS_SHEET = "SLA_Targets"
STAGE_FILTER_OPTIONS = ["Stage 1", "Stage 2", "Unmapped", "Missing Asset ID", "Needs Stage Review"]

# ── Stage text-detection (explicit "Stage 1" / "Stage 2" in work order fields) ─────────
# Require the word "stage" + digit so standalone severity codes "S1" / "S2" never match.
# Patterns cover: "Stage 2", "Stage2", "STAGE 2", "Stage-2", "Stage 2 Gemba Walk",
#                 "Gemba Walk Stage 2", etc.
_STAGE2_TEXT_RE = re.compile(r"\bstage[\s_-]*2\b", re.IGNORECASE)
_STAGE1_TEXT_RE = re.compile(r"\bstage[\s_-]*1\b", re.IGNORECASE)

# Fields scanned for explicit stage text.  Severity/priority fields are intentionally
# excluded to prevent "S2" severity from being misread as Stage 2.
_STAGE_TEXT_FIELDS = (
    "raw_functional_location",   # location code — most reliable stage signal
    "description",               # cleaned remarks / description text
    "location",                  # mapped location label
    "mapped_location",
    "mappedLocation",
    "machine_name_display",
    "machine_name",
    "raw_machine_name",
    "machine_group",
    "job_trade",
    "remarks",                   # raw remarks field if present on row
)

# ── Phase 3: SQL-backed enrichment helpers ──────��─────────────────────────────

# Machine groups whose category maps to Production / Utilities / Refrigeration
# (same classification used by asset_mapping.py).
_CRITICAL_CATEGORY_KEYWORDS = frozenset({
    "production", "refriger", "utilities", "utility",
})


def _sql_has_work_orders() -> bool:
    """True when the work_orders SQL table is populated (Phase 2 sync has run)."""
    try:
        import db as _db
        status = _db.get_db_status()
        return bool(status.get("ok")) and bool(status.get("work_orders_rows"))
    except Exception:
        return False


def _sql_asset_mapping_meta() -> dict:
    """Asset-mapping metadata from SQL so the SQL path never has to reopen Excel."""
    fallback = {
        "available": False,
        "path": None,
        "last_synced": None,
        "asset_count": 0,
        "keyword_rule_count": None,
        "group_count": 0,
        "message": "Asset master SQL metadata unavailable.",
        "data_source": "sql",
    }
    try:
        import db as _db
        meta = _db.get_asset_master_sync_meta()
        if isinstance(meta, dict):
            return meta
    except Exception as exc:
        fallback["message"] = f"Asset master SQL metadata unavailable: {exc}"
    return fallback


def _parse_iso_simple(s) -> datetime | None:
    """Parse an ISO-8601 string from SQL to a naive datetime, or return None."""
    if not s:
        return None
    text = str(s).strip().replace("Z", "+00:00")
    try:
        dt = datetime.fromisoformat(text)
        return dt.replace(tzinfo=None) if dt.tzinfo else dt
    except ValueError:
        return None


def _infer_criticality_from_category(category: str, machine_group: str) -> str:
    """Lightweight criticality inference from category/machine_group strings."""
    combined = (category + " " + machine_group).lower()
    for kw in _CRITICAL_CATEGORY_KEYWORDS:
        if kw in combined:
            return CRITICALITY_CRITICAL
    return CRITICALITY_NON_CRITICAL


def _sql_row_to_enriched(row: dict) -> dict:
    """
    Convert a raw SQL row (work_orders LEFT JOIN asset_master) to an
    enriched Python dict with the same field shape expected by
    build_management_downtime_payload() and the rest of the downtime stack.

    This is the Phase 3 equivalent of enrich_work_order_records() for a
    single already-enriched SQL row — it reconstructs derived fields from
    stored values rather than re-reading Excel.
    """
    mr_number    = str(row.get("mr_number")    or "").strip()
    wo_number    = str(row.get("wo_number")    or "").strip()
    asset_id     = str(row.get("asset_id")     or "").strip().upper()
    asset_name   = str(row.get("asset_name")   or "").strip()
    func_loc     = str(row.get("functional_location") or "").strip()
    stage        = str(row.get("stage")        or "").strip()
    category     = str(row.get("category")     or "").strip()
    machine_group = str(row.get("machine_group") or "").strip() or asset_name or asset_id or "Unmapped Asset"
    severity     = str(row.get("severity")     or "").strip()
    status       = str(row.get("status")       or "").strip()
    description  = str(row.get("description")  or "").strip()
    job_type     = str(row.get("job_type")     or "").strip()
    trade        = str(row.get("trade")        or "").strip()
    actual_start_iso  = row.get("actual_start")   or ""
    actual_end_iso    = row.get("actual_end")     or ""
    created_date_iso  = row.get("created_date")   or ""
    dq_status    = str(row.get("data_validity_status") or "").strip()
    review_reason = str(row.get("review_reason") or "").strip()

    # asset_master JOIN fields
    am_criticality = row.get("am_criticality")
    am_is_critical = row.get("am_is_critical")
    am_area        = str(row.get("am_area") or "").strip()
    has_am_match   = am_criticality is not None

    # Criticality
    if has_am_match and am_criticality:
        criticality = _normalize_criticality(am_criticality)
    else:
        criticality = _infer_criticality_from_category(category, machine_group)
    raw_criticality = am_criticality or ""
    is_critical_flag = bool(am_is_critical) if am_is_critical is not None else (criticality == CRITICALITY_CRITICAL)
    crit_rank = CRITICALITY_RANK.get(criticality, CRITICALITY_RANK.get("Unmapped", 2))

    location = am_area or "Unassigned"

    # Status flags
    status_lower = status.lower()
    is_open     = status_lower in {"new", "in progress", "inprogress"}
    is_finished = status_lower in {"finished", "completed", "closed", "resolved", "done"}

    is_valid = dq_status == "Valid"

    # Datetime parsing
    actual_start_dt = _parse_iso_simple(actual_start_iso)
    actual_end_dt   = _parse_iso_simple(actual_end_iso)
    created_dt      = _parse_iso_simple(created_date_iso)

    # Duration / TTR
    duration_hours = None
    ttr_source     = "excluded_status"
    duration_context = "Excluded from MTTR/TTR by lifecycle or data-quality rule"
    if is_valid and actual_start_dt and actual_end_dt and actual_end_dt > actual_start_dt:
        duration_hours = round((actual_end_dt - actual_start_dt).total_seconds() / 3600, 3)
        ttr_source     = "date_derived"
        duration_context = "Maintenance resolution time derived from valid Finished start/end dates"
    elif is_finished:
        ttr_source = "invalid_finished_dates"

    # start_time / end_time for period-overlap queries
    if is_valid and actual_start_dt:
        start_time = actual_start_dt.isoformat()
        end_time   = actual_end_dt.isoformat() if actual_end_dt else None
    else:
        start_time = created_dt.isoformat() if created_dt else None
        end_time   = None

    latest_ts = actual_end_dt or actual_start_dt or created_dt
    latest_event_time = latest_ts.isoformat() if latest_ts else None

    # Priority (numeric)
    priority = None
    try:
        pv = float(severity)
        if 1 <= pv <= 10:
            priority = int(pv)
    except (ValueError, TypeError):
        pass

    # Mapping metadata
    mapping_status = "Mapped" if has_am_match else "Unmapped"
    mapping_source = "Asset_Master.xlsx" if has_am_match else "fallback"

    # Data quality flags list
    if is_valid:
        dq_flags = ["Valid"]
    elif review_reason:
        dq_flags = [r.strip() for r in review_reason.split(";") if r.strip()]
    else:
        dq_flags = [dq_status or "Review"]

    # Acknowledgement
    if is_finished and wo_number:
        ack_status = "Acknowledged"
    elif is_open:
        ack_status = "Pending"
    else:
        ack_status = ""

    # status_category
    if is_open:
        status_category = "Open"
    elif is_finished:
        status_category = "Closed"
    else:
        status_category = "Review"

    return {
        # IDs
        "work_order_id": wo_number,
        "maintenance_order_id": mr_number,
        "asset_id": asset_id,
        # Asset / machine
        "asset_name": asset_name,
        "machine_group": machine_group,
        "machine_name": asset_name or machine_group,
        "machine_name_display": asset_name or machine_group,
        "machine_code": asset_id,
        "asset_display_name": asset_name or machine_group,
        "asset_label": asset_id,
        "raw_machine_name": asset_name,
        "machine_equipment_name": asset_name,
        # Location
        "location": location,
        "building": location,
        "area": location,
        # Stage
        "stage": stage,
        "resolved_stage": stage,
        "mapped_stage": stage,
        "mappedStage": stage,
        # Category / group
        "equipment_category": category,
        "mappedMainAssetGroup": category,
        "mapped_main_asset_group": category,
        "mappedSubAssetGroup": "",
        "mapped_sub_asset_group": "",
        "mappedLocation": location,
        "mapped_location": location,
        "mappedSystemArea": "",
        "mapped_system_area": "",
        "mappedAssetName": asset_name,
        "mapped_asset_name": asset_name,
        # Criticality
        "criticality": criticality,
        "raw_criticality": raw_criticality,
        "normalized_criticality": criticality,
        "is_critical": is_critical_flag,
        "criticality_rank": crit_rank,
        # Mapping metadata
        "mapping_status": mapping_status,
        "mappingStatus": mapping_status,
        "mapping_source": mapping_source,
        "classification_source": mapping_source,
        "has_assetlist_classification": has_am_match,
        "has_asset_master_mapping": has_am_match,
        "group_asset_ids": [asset_id] if asset_id else [],
        "refrigeration_group_match": machine_group == REFRIGERATION_GROUP,
        # Status
        "status": status,
        "request_state": status,
        "is_open": is_open,
        "status_category": status_category,
        # Dates
        "actual_start_time": actual_start_iso or None,
        "actual_end_time": actual_end_iso or None,
        "maintenance_start_time": actual_start_iso or None,
        "maintenance_end_time": actual_end_iso or None,
        "request_created_time": created_date_iso or None,
        "start_time": start_time,
        "end_time": end_time,
        "latest_event_time": latest_event_time,
        # TTR / duration
        "duration_hours": duration_hours,
        "ttr_hours": duration_hours,
        "raw_ttr": None,
        "ttr_source": ttr_source,
        "duration_context": duration_context,
        "valid_mttr_ttr": is_valid,
        # Data quality
        "data_quality_flag": dq_status or "Review",
        "data_quality_flags": dq_flags,
        # Description
        "description": description,
        "description_original": description,
        "translated_description": row.get("translated_description") or translate_maintenance_description(description),
        "remarks": description,
        # Job info
        "system": trade,
        "job_trade": trade,
        "maintenance_job_type": job_type,
        "raw_functional_location": func_loc,
        # Priority / severity
        "priority": priority,
        "service_level": severity,
        # Other
        "source": "Work Order",
        "source_path": str(row.get("source_file") or ""),
        "acknowledgement_status": ack_status,
        "started_by": str(row.get("started_by") or ""),
        "created_by": str(row.get("created_by") or ""),
    }


def load_work_order_downtime_sql(stage: str | None = None) -> dict:
    """
    SQL-backed replacement for load_work_order_downtime().

    Reads enriched records directly from the work_orders SQL table (populated by
    Phase 2 after each file import).  Stage filtering is pushed into the SQL query,
    so callers do NOT need to run filter_work_orders_by_stage().

    Returns the same dict shape as load_work_order_downtime():
        {"available": bool, "records": [...], "last_synced": str|None, "message": str}
    """
    normalized_stage = normalize_stage_filter(stage)
    cache_key = normalized_stage or "__all__"
    cached = _SQL_WO_CACHE.get(cache_key)
    if cached is not None:
        return cached

    try:
        import db as _db
        sql_rows = _db.load_work_orders_from_sql(normalized_stage or None)
    except Exception as exc:
        return {
            "available": False,
            "records": [],
            "message": f"SQL query failed: {exc}",
            "last_synced": None,
        }

    if not sql_rows:
        return {
            "available": False,
            "records": [],
            "message": "No work order data found in SQL database.",
            "last_synced": None,
        }

    records = [_sql_row_to_enriched(r) for r in sql_rows]

    # Use the most recent updated_at as the "last synced" timestamp.
    # Strip trailing "Z" so downstream pd.Timestamp() gets a naive datetime,
    # matching the format produced by the Excel-based loader.
    updated_ats = [r.get("updated_at") for r in sql_rows if r.get("updated_at")]
    last_synced = max(updated_ats).rstrip("Z") if updated_ats else None

    result = {
        "available": True,
        "records": records,
        "message": f"Loaded {len(records)} work order(s) from SQL.",
        "last_synced": last_synced,
    }
    _SQL_WO_CACHE[cache_key] = result
    return result


DEFAULT_SLA_TARGETS = [
    {
        "key": "S1",
        "label": "S1 Critical",
        "short_label": "S1",
        "fallback_severity": "Critical",
        "response_target_hours": 1,
        "completion_target_hours": None,
        "rank": 1,
    },
    {
        "key": "S2",
        "label": "S2 High",
        "short_label": "S2",
        "fallback_severity": "High",
        "response_target_hours": 4,
        "completion_target_hours": 72,
        "rank": 2,
    },
    {
        "key": "S3",
        "label": "S3 Medium",
        "short_label": "S3",
        "fallback_severity": "Medium",
        "response_target_hours": 48,
        "completion_target_hours": 504,
        "rank": 3,
    },
    {
        "key": "S4",
        "label": "S4 Low",
        "short_label": "S4",
        "fallback_severity": "Low",
        "response_target_hours": None,
        "completion_target_hours": 1080,
        "rank": 4,
    },
]

SLA_TARGET_HEADER_ALIASES = {
    "key": {"key", "severitykey", "servicelevel", "servicelevelkey", "severity"},
    "label": {"label", "severitylabel", "servicelevellabel", "displaylabel"},
    "short_label": {"shortlabel", "short", "badge", "code"},
    "fallback_severity": {"fallbackseverity", "severityname", "prioritylabel", "mappedseverity"},
    "response_target_hours": {"responsetargethours", "responsehours", "responsehrs", "acktargethours"},
    "completion_target_hours": {"completiontargethours", "completionhours", "completionhrs", "closetargethours"},
    "rank": {"rank", "sort", "sortorder", "order"},
    "active": {"active", "enabled", "include"},
}

WORK_ORDER_COLUMN_ALIASES = {
    "Request State": ["Request State", "Status", "WO Status", "Work Order Status", "Request Status", "Current lifecycle state", "Current lifecycle state2", "Current lifecycle state3", "Lifecycle State", "request_state"],
    "Request ID": ["Request ID", "Request No", "Request Number", "Maintenance Request", "Maintenance Request ID", "MNT ID", "maintenance_order_id"],
    "WO ID": ["WO ID", "Work Order ID", "WorkOrder ID", "Work Order", "Work order", "Work Order No", "WO No", "WO Number", "work_order_id"],
    "Machine ID": ["Machine ID", "Machine Code", "Asset", "Asset ID", "AssetID", "Equipment ID", "Equipment Code", "machine_code"],
    "Machine Name": ["Machine Name", "Name", "Asset Name", "Equipment Name", "Machine", "Asset Description", "machine_name"],
    "Description": ["Description", "Work Description", "Problem Description", "Job Description", "Notes", "Remarks"],
    "Location": ["Location", "Building", "Area", "Production Line", "Work Area", "Functional location"],
    "JobTrade": ["JobTrade", "Job Trade", "Trade", "System", "Work Type", "Maintenance Type", "Job Category"],
    "JobTypeId": ["JobTypeId", "Job Type ID", "Job Type", "JobType", "Maintenance job type", "Maintenance job type variant", "Maintenance request type"],
    "Priority": ["Priority", "Priority No", "Priority Number", "Priority Level", "Severity", "SeverityLevel", "Service level", "severity"],
    "Started By": ["Started By", "Started by", "started_by", "StartedBy"],
    "Created By": ["Created By", "Created by", "created_by", "CreatedBy"],
    "Request Created Date": ["Request Created Date", "Created date and time", "Created Date", "Created On", "Reported Date", "Request Date"],
    "Actual Start Date": ["Actual Start Date", "Actual Start", "Actual start", "Maintenance Start Date", "Start Date"],
    "Actual End Date": ["Actual End Date", "Actual End", "Actual end", "Maintenance End Date", "End Date", "Closed Date", "Completed Date"],
    "TTR(hr)": ["TTR(hr)", "TTR", "TTR Hours", "TTR Hour", "Time to Resolution", "Resolution Time", "Downtime Hours", "Duration Hours", "downtime_hours"],
    "TTR Minutes": ["TTR Minutes", "TTR(min)", "TTR Minute", "TTR Mins", "Duration Minutes", "Downtime Minutes"],
}
WORK_ORDER_REQUIRED_CANONICAL_COLUMNS = {
    "Request State", "Request ID", "WO ID", "Machine ID", "Machine Name",
    "Priority", "Actual Start Date", "Actual End Date", "TTR(hr)",
}


def normalize_period(value):
    cleaned = str(value or "").strip().lower()
    if cleaned in {"this_month", "this month", "current month", "mtd", "month to date", "monthtodate"}:
        return "this_month"
    if cleaned in {"last_month", "last month", "previous month"}:
        return "last_month"
    if cleaned in {"7d", "week", "next week", "last 7 days"}:
        return "7d"
    if cleaned in {"last12", "last_12_months", "last 12 months", "12m", "rolling12"}:
        return "last12"
    if cleaned in {"previous_year", "previous year", "last year"}:
        return "previous_year"
    if cleaned in {"custom", "custom_range", "custom date range"}:
        return "custom"
    if cleaned in {"all_years", "all years", "all", "historical"}:
        return "all_years"
    if cleaned in {"90d", "quarter", "qtr", "last 90 days"}:
        return "90d"
    if cleaned in {"ytd", "year", "year to date", "full year"}:
        return "ytd"
    if cleaned in {"30d", "last 30 days"}:
        return "30d"
    return "ytd"


def normalize_stage_filter(value):
    cleaned = str(value or "").strip()
    key = re.sub(r"[^a-z0-9]+", "", cleaned.lower())
    if not key or key in {"all", "allstages"}:
        return ""
    if key in {"stage1", "st1", "s1"}:
        return "Stage 1"
    if key in {"stage2", "st2", "s2"}:
        return "Stage 2"
    if key in {"unmapped"}:
        return "Unmapped"
    if key in {"missingassetid", "missingasset", "missingid"}:
        return "Missing Asset ID"
    if key in {"needsstagereview", "stagereview", "review", "keywordmatched"}:
        return "Needs Stage Review"
    return cleaned if cleaned in STAGE_FILTER_OPTIONS else ""


def get_asset_master_path(data_dir=DATA_DIR):
    preferred = os.path.join(data_dir, ASSET_MASTER_RELATIVE_PATH)
    fallback = os.path.join(data_dir, ASSET_MASTER_FILENAME)
    if os.path.exists(preferred) or not os.path.exists(fallback):
        return preferred
    return fallback


def _asset_mapped_stage(row):
    """Read the asset-master-assigned stage from an enriched record (no text scanning)."""
    status = str(row.get("mappingStatus") or row.get("mapping_status") or "").strip()
    stage = str(row.get("mappedStage") or row.get("mapped_stage") or "").strip()
    if status in {"Unmapped", "Missing Asset ID"}:
        return status
    if stage in {"Stage 1", "Stage 2", "Needs Stage Review"}:
        return stage
    if status in {"Needs Stage Review", "Keyword Matched"}:
        return "Needs Stage Review"
    return stage or status or "Unmapped"


def detect_stage_from_text(row):
    """Return "Stage 1" or "Stage 2" if the MR/WO fields explicitly state a stage; else None.

    Scans description, location, functional location, machine name, and related text
    fields.  Severity/priority fields are deliberately excluded so that an "S2 severity"
    code is never mistaken for Stage 2.  Only the word "stage" followed by 1 or 2 (with
    optional separator) is accepted — e.g. "Stage 2 Gemba Walk", "Gemba Walk Stage 2",
    "Stage-1", "Stage1", "STAGE 2".
    """
    combined = " ".join(str(row.get(f) or "") for f in _STAGE_TEXT_FIELDS if row.get(f))
    if not combined.strip():
        return None
    # Check Stage 2 first — a record that explicitly says Stage 2 must not appear under Stage 1.
    if _STAGE2_TEXT_RE.search(combined):
        return "Stage 2"
    if _STAGE1_TEXT_RE.search(combined):
        return "Stage 1"
    return None


def resolve_work_order_stage(row):
    """Central stage resolver — single source of truth for all downtime stage classification.

    Priority:
      1. Explicit stage text in description, location, or functional location fields.
         Text detection takes precedence over asset mapping when the text clearly names
         a stage (e.g. "Stage 2 Gemba Walk" → Stage 2 even if asset is mapped to Stage 1).
      2. mappedStage from the Asset Master classification.
      3. Existing mapping status (Unmapped / Missing Asset ID / Needs Stage Review).

    Debug logging (Python DEBUG level on this module) shows:
      - resolved stage and reason (text-detected vs asset-mapping)
      - original asset_id, mappedStage, mapping status, and the text fields that triggered
        the decision — useful for validating classification without changing production output.
    """
    text_stage = detect_stage_from_text(row)
    if text_stage is not None:
        _log.debug(
            "resolve_stage=%s [text-detected] | asset_id=%s | mappedStage=%s | text_fields=%s",
            text_stage,
            row.get("asset_id") or row.get("machine_code"),
            row.get("mappedStage"),
            {f: row[f] for f in _STAGE_TEXT_FIELDS if row.get(f)},
        )
        return text_stage

    asset_stage = _asset_mapped_stage(row)
    _log.debug(
        "resolve_stage=%s [asset-mapping] | asset_id=%s | mappedStage=%s | status=%s",
        asset_stage,
        row.get("asset_id") or row.get("machine_code"),
        row.get("mappedStage"),
        row.get("mappingStatus"),
    )
    return asset_stage


def get_work_order_stage_scope(row):
    """Return the resolved stage for a work order / MR record.

    Delegates to resolve_work_order_stage() — the single authoritative resolver
    that combines text-based stage detection with Asset Master classification.
    All filtering and aggregation must go through this function.
    """
    return resolve_work_order_stage(row)


def filter_work_orders_by_stage(records, stage_filter):
    normalized_stage = normalize_stage_filter(stage_filter)
    if not normalized_stage:
        return list(records or [])
    # Use pre-computed resolved_stage when available (set by load_work_order_downtime);
    # fall back to live resolution so callers that skip the annotator still work.
    return [
        row for row in records or []
        if (row.get("resolved_stage") if row.get("resolved_stage") is not None
            else get_work_order_stage_scope(row)) == normalized_stage
    ]


def get_period_days(period):
    return {"7d": 7, "30d": 30, "90d": 90, "last12": 365}.get(period, 30)


def get_period_label(period):
    return {
        "this_month": "This Month",
        "last_month": "Last Month",
        "7d": "Last 7 Days",
        "30d": "Last 30 Days",
        "90d": "Quarter",
        "ytd": "Current Year / YTD",
        "last12": "Last 12 Months",
        "previous_year": "Previous Year",
        "custom": "Custom Date Range",
        "all_years": "All Years",
    }.get(period, "Current Year / YTD")


def normalize_date_filter(value):
    parsed = pd.to_datetime(str(value or "").strip(), errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.to_pydatetime().replace(hour=0, minute=0, second=0, microsecond=0)


def normalize_month_filter(value):
    cleaned = str(value or "").strip()
    if not cleaned or cleaned.lower() == "all":
        return None
    parsed = pd.to_datetime(f"{cleaned}-01", errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.strftime("%Y-%m")


def normalize_config_header(value):
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def clean_config_text(value, fallback=""):
    if value is None or pd.isna(value):
        return fallback
    text = re.sub(r"\s+", " ", str(value).strip())
    return text or fallback


def parse_optional_hours(value):
    if value is None or pd.isna(value):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value) if float(value) % 1 else int(value)
    text = str(value).strip().lower()
    if not text or text in {"-", "--", "n/a", "na", "none", "null", "blank", "no target"}:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text.replace(",", ""))
    if not match:
        return None
    amount = float(match.group(0))
    if "day" in text or re.search(r"\bd\b", text):
        amount *= 24
    elif "min" in text:
        amount /= 60
    return amount if amount % 1 else int(amount)


def parse_config_rank(value, fallback):
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return fallback


def parse_config_active(value):
    if value is None or pd.isna(value):
        return True
    return str(value).strip().lower() not in {"0", "false", "no", "n", "inactive", "disabled"}


def get_sla_header_key(raw_header):
    normalized = normalize_config_header(raw_header)
    for key, aliases in SLA_TARGET_HEADER_ALIASES.items():
        if normalized in aliases:
            return key
    return None


def copy_default_sla_targets():
    return [dict(target) for target in DEFAULT_SLA_TARGETS]


def _default_sla_target_config():
    return {
        "available": False,
        "source": f"{ASSET_MASTER_RELATIVE_PATH}:{SLA_TARGETS_SHEET}",
        "message": (
            f"Using built-in default SLA targets without reopening {ASSET_MASTER_FILENAME} "
            "during SQL-backed overview loading."
        ),
        "targets": copy_default_sla_targets(),
        "instructions": (
            "Edit Response Target Hours and Completion Target Hours in the "
            f"'{SLA_TARGETS_SHEET}' sheet to change these defaults."
        ),
    }


def load_sla_target_config(data_dir=DATA_DIR):
    path = get_asset_master_path(data_dir)
    defaults = copy_default_sla_targets()
    fallback = {
        "available": False,
        "source": f"{ASSET_MASTER_RELATIVE_PATH}:{SLA_TARGETS_SHEET}",
        "message": f"Using built-in default SLA targets. Add or edit the '{SLA_TARGETS_SHEET}' sheet in {ASSET_MASTER_FILENAME} to change them.",
        "targets": defaults,
        "instructions": "Edit Response Target Hours and Completion Target Hours. Leave a target cell blank for no target; set Active to FALSE to remove both targets for that severity.",
    }
    if not os.path.exists(path):
        fallback["message"] = f"{ASSET_MASTER_RELATIVE_PATH} not found; using built-in default SLA targets."
        return fallback

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        fallback["message"] = f"Could not read {ASSET_MASTER_FILENAME}; using built-in default SLA targets. {exc}"
        return fallback

    try:
        if SLA_TARGETS_SHEET not in wb.sheetnames:
            return fallback

        ws = wb[SLA_TARGETS_SHEET]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        try:
            wb.close()
        except Exception:
            pass
    if not rows:
        fallback["message"] = f"'{SLA_TARGETS_SHEET}' is empty; using built-in default SLA targets."
        return fallback

    header_row_index = None
    header_map = {}
    for index, row in enumerate(rows[:10]):
        mapped = {get_sla_header_key(cell): offset for offset, cell in enumerate(row) if get_sla_header_key(cell)}
        if "key" in mapped:
            header_row_index = index
            header_map = mapped
            break

    if header_row_index is None:
        fallback["message"] = f"'{SLA_TARGETS_SHEET}' has no Severity Key/Service Level column; using built-in default SLA targets."
        return fallback

    defaults_by_key = {target["key"]: dict(target) for target in defaults}
    loaded_by_key = {}
    for row in rows[header_row_index + 1:]:
        key = clean_config_text(row[header_map["key"]] if header_map.get("key", -1) < len(row) else "").upper()
        if not key or key.startswith("#"):
            continue
        base = defaults_by_key.get(key, {
            "key": key,
            "label": key,
            "short_label": key,
            "fallback_severity": key,
            "response_target_hours": None,
            "completion_target_hours": None,
            "rank": 99,
        })
        target = dict(base)
        is_active = "active" not in header_map or parse_config_active(row[header_map["active"]] if header_map["active"] < len(row) else None)
        if "label" in header_map:
            target["label"] = clean_config_text(row[header_map["label"]] if header_map["label"] < len(row) else None, target["label"])
        if "short_label" in header_map:
            target["short_label"] = clean_config_text(row[header_map["short_label"]] if header_map["short_label"] < len(row) else None, target["short_label"])
        if "fallback_severity" in header_map:
            target["fallback_severity"] = clean_config_text(row[header_map["fallback_severity"]] if header_map["fallback_severity"] < len(row) else None, target["fallback_severity"])
        if not is_active:
            target["response_target_hours"] = None
            target["completion_target_hours"] = None
        elif "response_target_hours" in header_map:
            target["response_target_hours"] = parse_optional_hours(row[header_map["response_target_hours"]] if header_map["response_target_hours"] < len(row) else None)
        if is_active and "completion_target_hours" in header_map:
            target["completion_target_hours"] = parse_optional_hours(row[header_map["completion_target_hours"]] if header_map["completion_target_hours"] < len(row) else None)
        if "rank" in header_map:
            target["rank"] = parse_config_rank(row[header_map["rank"]] if header_map["rank"] < len(row) else None, target["rank"])
        loaded_by_key[key] = target

    if not loaded_by_key:
        fallback["message"] = f"'{SLA_TARGETS_SHEET}' contains no readable SLA target rows; using built-in default SLA targets."
        return fallback

    for key, target in defaults_by_key.items():
        loaded_by_key.setdefault(key, target)

    return {
        "available": True,
        "source": f"{ASSET_MASTER_RELATIVE_PATH}:{SLA_TARGETS_SHEET}",
        "message": f"SLA targets loaded from {ASSET_MASTER_FILENAME} sheet '{SLA_TARGETS_SHEET}'.",
        "targets": sorted(loaded_by_key.values(), key=lambda target: (target.get("rank", 99), target.get("key", ""))),
        "instructions": "Edit Response Target Hours and Completion Target Hours. Leave a target cell blank for no target; set Active to FALSE to remove both targets for that severity.",
    }


def format_month_label(month_key):
    parsed = pd.to_datetime(f"{month_key}-01", errors="coerce")
    if pd.isna(parsed):
        return month_key
    return parsed.strftime("%b %Y")


def build_year_month_options(reference_dt):
    if reference_dt is None:
        return []
    year = reference_dt.year
    current_month = reference_dt.month
    return [f"{year}-{month:02d}" for month in range(1, current_month + 1)]


def build_year_month_options_from_events(events, reference_dt):
    timestamps = []
    for event in events or []:
        parsed = parse_iso_datetime(event.get("start_time") or event.get("end_time"))
        if parsed is not None:
            timestamps.append(parsed)
    if not timestamps:
        return build_year_month_options(reference_dt)
    start = min(timestamps).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    end = (reference_dt.to_pydatetime() if hasattr(reference_dt, "to_pydatetime") else reference_dt) or max(timestamps)
    end = end.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    values = []
    current = start
    while current <= end:
        values.append(f"{current.year}-{current.month:02d}")
        if current.month == 12:
            current = current.replace(year=current.year + 1, month=1)
        else:
            current = current.replace(month=current.month + 1)
    return values


def parse_iso_datetime(value):
    if value is None or pd.isna(value):
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None) if value.tzinfo else value
    text = str(value).strip()
    if not text:
        return None
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        return parsed.replace(tzinfo=None) if parsed.tzinfo else parsed
    except ValueError:
        parsed = pd.to_datetime(text, errors="coerce")
        if pd.isna(parsed):
            return None
        parsed_dt = parsed.to_pydatetime()
        return parsed_dt.replace(tzinfo=None) if parsed_dt.tzinfo else parsed_dt


def infer_default_year_from_path(path):
    match = re.search(r"(20\d{2}|19\d{2})", os.path.basename(str(path)))
    if match:
        return int(match.group(1))
    return DOWNTIME_EXPORT_YEAR


def get_work_order_source_paths():
    imported_candidates = []
    if os.path.isdir(WORK_ORDER_IMPORT_DIR):
        for extension in WORK_ORDER_IMPORT_EXTENSIONS:
            imported_candidates.extend(glob(os.path.join(WORK_ORDER_IMPORT_DIR, f"*{extension}")))

    if imported_candidates:
        candidates = sorted(imported_candidates, key=lambda path: os.path.getmtime(path), reverse=True)
    else:
        candidates = []
        data_patterns = [
            os.path.join(DATA_DIR, "data downtime.csv"),
            os.path.join(DATA_DIR, "data downtime *.csv"),
            os.path.join(DATA_DIR, "data_downtime*.csv"),
        ]
        for pattern in data_patterns:
            candidates.extend(glob(pattern))
        if os.path.exists(FALLBACK_WORK_ORDER_DOWNTIME_FILE):
            candidates.append(FALLBACK_WORK_ORDER_DOWNTIME_FILE)
    ordered = []
    seen = set()
    for path in candidates:
        normalized = os.path.abspath(path)
        if normalized in seen or not os.path.exists(normalized):
            continue
        seen.add(normalized)
        ordered.append(normalized)
    return ordered


def read_work_order_source_file(path):
    extension = os.path.splitext(path)[1].lower()
    if extension == ".csv":
        df = pd.read_csv(path, encoding="utf-8-sig")
    elif extension in {".xlsx", ".xls"}:
        sheets = pd.read_excel(path, sheet_name=None)
        if not sheets:
            raise ValueError("Workbook contains no readable sheets.")
        sheet_name, df = max(sheets.items(), key=lambda item: score_work_order_dataframe(item[1]))
        if score_work_order_dataframe(df) == 0:
            raise ValueError("Workbook does not contain recognizable work order columns.")
    else:
        raise ValueError(f"Unsupported work order import file type: {extension}")
    return canonicalize_work_order_dataframe(df)


def normalize_work_order_column_key(value):
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def build_work_order_alias_lookup():
    lookup = {}
    for canonical, aliases in WORK_ORDER_COLUMN_ALIASES.items():
        lookup[normalize_work_order_column_key(canonical)] = canonical
        for alias in aliases:
            lookup[normalize_work_order_column_key(alias)] = canonical
    return lookup


WORK_ORDER_ALIAS_LOOKUP = build_work_order_alias_lookup()


def score_work_order_dataframe(df):
    columns = {
        WORK_ORDER_ALIAS_LOOKUP.get(normalize_work_order_column_key(column), str(column).strip())
        for column in getattr(df, "columns", [])
    }
    return sum(1 for column in WORK_ORDER_REQUIRED_CANONICAL_COLUMNS if column in columns)


def canonicalize_work_order_dataframe(df):
    df = df.copy()
    renamed_columns = []
    seen = set()
    for column in df.columns:
        original = str(column).replace("\ufeff", "").strip()
        canonical = WORK_ORDER_ALIAS_LOOKUP.get(normalize_work_order_column_key(original), original)
        # Keep duplicate aliases under their original names so no source column is silently overwritten.
        if canonical in seen and canonical != original:
            canonical = original
        seen.add(canonical)
        renamed_columns.append(canonical)
    df.columns = renamed_columns
    unnamed = [column for column in df.columns if str(column).lower().startswith("unnamed:") and df[column].isna().all()]
    if unnamed:
        df = df.drop(columns=unnamed)
    return df


def get_work_order_import_status():
    paths = get_work_order_source_paths()
    import_root = os.path.abspath(WORK_ORDER_IMPORT_DIR)
    return {
        "source_count": len(paths),
        "using_uploaded_imports": bool(paths) and all(os.path.abspath(path).startswith(import_root) for path in paths),
        "sources": [
            {
                "name": os.path.basename(path),
                "path": path,
                "last_modified": datetime.fromtimestamp(os.path.getmtime(path)).isoformat(),
                "size": os.path.getsize(path),
            }
            for path in paths
        ],
    }


def build_mtbf_work_order_history_payload(stage=None):
    if _sql_has_work_orders():
        payload = load_work_order_downtime_sql(stage)
        source_rows = payload.get("records") or []
    else:
        payload = load_work_order_downtime()
        source_rows = filter_work_orders_by_stage(payload.get("records") or [], stage)
    records = []
    years = set()
    months = set()

    for row in source_rows:
        start_time = row.get("actual_start_time") or row.get("maintenance_start_time") or row.get("start_time")
        end_time = row.get("actual_end_time") or row.get("maintenance_end_time") or row.get("end_time")
        start_dt = parse_iso_datetime(start_time)
        if start_dt is not None:
            years.add(str(start_dt.year))
            months.add(f"{start_dt.year}-{start_dt.month:02d}")

        records.append(
            {
                "asset_id": row.get("asset_id"),
                "machine_group": row.get("machine_group") or row.get("machine_name_display") or row.get("machine_name"),
                "criticality": row.get("criticality"),
                "raw_criticality": row.get("raw_criticality"),
                "start_time": start_time,
                "end_time": end_time,
                "actual_start_time": row.get("actual_start_time"),
                "actual_end_time": row.get("actual_end_time"),
                "work_order_id": row.get("work_order_id"),
            }
        )

    return {
        "available": payload.get("available", False),
        "last_synced": payload.get("last_synced"),
        "years": sorted(years, reverse=True),
        "months": sorted(months, reverse=True),
        "work_order_count": len(records),
        "work_orders": records,
        "sources": get_work_order_import_status().get("sources", []),
    }


def get_first_present(row, *names):
    for name in names:
        if name in row and not pd.isna(row.get(name)):
            value = row.get(name)
            if str(value).strip():
                return value
    return None


def has_present_value(value):
    if value is None:
        return False
    try:
        if pd.isna(value):
            return False
    except (TypeError, ValueError):
        pass
    return bool(str(value).strip())


def parse_work_order_datetime(value, end_of_day=False):
    if value is None or pd.isna(value):
        return None
    parsed = pd.to_datetime(value, errors="coerce", dayfirst=False)
    if pd.isna(parsed):
        parsed = pd.to_datetime(value, errors="coerce", dayfirst=True)
    if pd.isna(parsed):
        return None
    dt = parsed.to_pydatetime()
    if end_of_day and dt.hour == 0 and dt.minute == 0 and dt.second == 0 and dt.microsecond == 0:
        dt = dt.replace(hour=23, minute=59, second=59)
    return dt


def parse_work_order_datetime_with_quality(value, end_of_day=False):
    if not has_present_value(value):
        return None, False
    parsed = parse_work_order_datetime(value, end_of_day=end_of_day)
    return parsed, parsed is None


def normalize_lifecycle_state(value):
    return re.sub(r"\s+", " ", clean_ascii_text(value).strip().lower())


def is_new_lifecycle(value):
    return normalize_lifecycle_state(value) == "new"


def is_in_progress_lifecycle(value):
    return normalize_lifecycle_state(value).replace(" ", "") == "inprogress"


def is_finished_lifecycle(value):
    return normalize_lifecycle_state(value) == "finished"


def is_review_lifecycle(value):
    normalized = normalize_lifecycle_state(value)
    return normalized in {"confirm", "rework", "re work", "rejected", "reject"}


def calculate_acknowledgement_status(status, work_order_id):
    has_work_order = has_present_value(work_order_id)
    if is_new_lifecycle(status) and not has_work_order:
        return "Not Acknowledged"
    if is_in_progress_lifecycle(status) and has_work_order:
        return "Acknowledged / In Progress"
    if is_finished_lifecycle(status):
        return "Closed"
    return "Review"


def build_work_order_quality_flags(
    status,
    request_created_raw,
    request_created_time,
    request_created_invalid,
    actual_start_raw,
    actual_start_time,
    actual_start_invalid,
    actual_end_raw,
    actual_end_time,
    actual_end_invalid,
):
    # Dynamics lifecycle rules keep acknowledgement, TTR, and reliability history separate.
    flags = []
    if request_created_invalid or actual_start_invalid or actual_end_invalid:
        flags.append("Invalid date format")
    if not request_created_time:
        flags.append("Missing raised date")
    if is_finished_lifecycle(status):
        if not actual_start_time:
            flags.append("Missing start date for finished MR")
        if not actual_end_time:
            flags.append("Missing finished date for finished MR")
        if actual_start_time and actual_end_time and actual_end_time < actual_start_time:
            flags.append("Finished date before start date")
        if request_created_time and actual_end_time and actual_end_time < request_created_time:
            flags.append("Finished date before raised date")
    elif is_new_lifecycle(status) or is_in_progress_lifecycle(status):
        if has_present_value(actual_end_raw) and actual_end_time:
            flags.append("Unexpected finished date for New/In Progress MR")
    else:
        flags.append("Review status")
    return flags or ["Valid"]


def write_work_orders_to_db() -> dict:
    """
    Load all enriched work-order records (via the in-process cache) and upsert
    them into the SQLite work_orders table.  Also writes one row to import_log.
    Called from a background thread after a successful file import.
    """
    import db as _db
    payload = load_work_order_downtime()
    if not payload.get("available"):
        return {"ok": False, "message": payload.get("message", "No data available.")}
    records = payload.get("records") or []
    source_file = ""
    if records:
        source_file = os.path.basename(records[0].get("source_path") or "")
    result = _db.upsert_work_orders(records, source_file)
    _db.log_import(
        source_type="work_orders",
        source_file=source_file,
        row_count=len(records),
        valid_count=result["valid"],
        invalid_count=result["invalid"],
        notes="Auto-sync after import",
    )
    return {
        "ok": True,
        "rows": result["rows"],
        "valid": result["valid"],
        "invalid": result["invalid"],
        "message": f"Synced {result['rows']} work order(s) into SQL ({result['valid']} valid, {result['invalid']} review).",
    }


def _write_wo_to_db_background():
    try:
        result = write_work_orders_to_db()
        print(f"[db] {result['message']}")
    except Exception as exc:
        print(f"[db] work_orders sync error: {exc}")


def import_work_order_file(file_storage, replace=True):
    filename = os.path.basename(getattr(file_storage, "filename", "") or "")
    extension = os.path.splitext(filename)[1].lower()
    if extension not in WORK_ORDER_IMPORT_EXTENSIONS:
        return {
            "ok": False,
            "message": "Unsupported file type. Upload a CSV, XLSX, or XLS work order export.",
        }

    os.makedirs(WORK_ORDER_IMPORT_DIR, exist_ok=True)
    safe_base = re.sub(r"[^A-Za-z0-9_.-]+", "_", os.path.splitext(filename)[0]).strip("._") or "work_orders"
    target_name = f"{safe_base}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{extension}"
    target_path = os.path.join(WORK_ORDER_IMPORT_DIR, target_name)
    file_storage.save(target_path)

    try:
        df = read_work_order_source_file(target_path)
    except Exception as exc:
        try:
            os.remove(target_path)
        except OSError:
            pass
        return {
            "ok": False,
            "message": f"Work order file could not be read: {exc}",
        }

    if replace:
        import_root = os.path.abspath(WORK_ORDER_IMPORT_DIR)
        for existing in get_work_order_source_paths():
            existing_path = os.path.abspath(existing)
            if existing_path == os.path.abspath(target_path):
                continue
            if existing_path.startswith(import_root):
                try:
                    os.remove(existing)
                except OSError:
                    pass

    _DOWNTIME_CACHE.clear()
    _WO_LOAD_CACHE["sig"] = None
    _WO_LOAD_CACHE["payload"] = None
    _SQL_WO_CACHE.clear()

    threading.Thread(target=_write_wo_to_db_background, name="db-wo-sync", daemon=True).start()

    return {
        "ok": True,
        "message": f"Imported {len(df)} work order row(s).",
        "file": target_name,
        "rows": int(len(df)),
        "columns": [str(col) for col in df.columns],
    }


def normalize_key(value):
    return "".join(ch for ch in str(value).strip().lower() if ch.isalnum())


def has_thai_text(value):
    return bool(re.search(r"[\u0E00-\u0E7F]", str(value or "")))


def thai_char_count(value):
    return len(re.findall(r"[\u0E00-\u0E7F]", str(value or "")))


def clean_ascii_text(value, fallback=""):
    if value is None:
        return fallback
    try:
        if pd.isna(value):
            return fallback
    except (TypeError, ValueError):
        pass
    text = str(value or "").replace("\ufeff", "").strip()
    text = text.replace("Producton", "Production").replace("producton", "Production")
    text = text.replace("Buiding", "Building").replace("buiding", "Building")
    text = re.sub(r"[^\x20-\x7E]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip(" -_/")
    return text or fallback


def clean_unicode_text(value, fallback=""):
    if value is None:
        return fallback
    try:
        if pd.isna(value):
            return fallback
    except (TypeError, ValueError):
        pass
    text = str(value or "").replace("\ufeff", " ").strip()
    text = text.replace("Producton", "Production").replace("producton", "Production")
    text = text.replace("Buiding", "Building").replace("buiding", "Building")
    text = re.sub(r"\s+", " ", text).strip(" -_/")
    return text or fallback


_TRANSLATION_CACHE_FILE = os.path.abspath(os.path.join(BASE_DIR, "..", "data", "translation_cache.json"))
_translation_cache: dict[str, str] = {}
_translation_cache_dirty = False
_translation_cache_mtime = None
_translation_lock = threading.Lock()
_translation_pending: set = set()
_translation_worker_active = False
_argos_translate_module = None
# Clear downtime payload cache every N background translations so browsers pick up new translations.
_BG_TRANSLATE_CLEAR_EVERY = 200


def _load_translation_cache():
    global _translation_cache, _translation_cache_mtime
    for path in (_TRANSLATION_CACHE_FILE, _TRANSLATION_CACHE_FILE + ".tmp"):
        try:
            if os.path.exists(path):
                with open(path, encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, dict) and data:
                    _translation_cache = data
                    _translation_cache_mtime = os.path.getmtime(path)
                    return
        except Exception:
            pass
    _translation_cache = {}
    _translation_cache_mtime = None


def _save_translation_cache():
    global _translation_cache_dirty
    if not _translation_cache_dirty:
        return
    try:
        with _translation_lock:
            snapshot = dict(_translation_cache)
            _translation_cache_dirty = False
        tmp_path = _TRANSLATION_CACHE_FILE + ".tmp"
        with open(tmp_path, "w", encoding="utf-8") as f:
            json.dump(snapshot, f, ensure_ascii=False, indent=2)
        os.replace(tmp_path, _TRANSLATION_CACHE_FILE)
    except Exception:
        pass


_load_translation_cache()


def _refresh_translation_cache_if_changed():
    """Pick up translations saved by the background worker without requiring a backend restart."""
    try:
        if not os.path.exists(_TRANSLATION_CACHE_FILE):
            return
        mtime = os.path.getmtime(_TRANSLATION_CACHE_FILE)
        if _translation_cache_mtime is None or mtime > _translation_cache_mtime:
            _load_translation_cache()
    except Exception:
        pass


def _clean_translation_output(value: str) -> str:
    text = clean_unicode_text(value)
    text = re.sub(r"\bAir\s*blast\s*(\d+)", r"Air Blast \1", text, flags=re.IGNORECASE)
    text = re.sub(r"\bAir\s*blast\s+one\b", "Air Blast 1", text, flags=re.IGNORECASE)
    text = re.sub(r"\bBratt\s*Pan\s*(\d+)", r"Bratt Pan \1", text, flags=re.IGNORECASE)
    text = re.sub(r"\bCold\s*(\d+)", r"Cold Room \1", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+([,./)])", r"\1", text)
    text = re.sub(r"([(])\s+", r"\1", text)
    return re.sub(r"\s+", " ", text).strip()


def _translate_with_argos(text: str) -> str:
    """Offline Thai-to-English fallback using the downloaded Argos model."""
    global _argos_translate_module
    if _argos_translate_module is False:
        return ""
    try:
        if _argos_translate_module is None:
            import argostranslate.translate as argos_translate
            _argos_translate_module = argos_translate
        translated = _argos_translate_module.translate(text, "th", "en")
        return _clean_translation_output(translated)
    except Exception:
        _argos_translate_module = False
        return ""


def _bg_translation_worker():
    """Daemon thread: drains _translation_pending using the Google Translate API."""
    global _translation_worker_active, _translation_cache_dirty
    completed = 0
    while True:
        with _translation_lock:
            if not _translation_pending:
                _translation_worker_active = False
                break
            text = next(iter(_translation_pending))
            _translation_pending.discard(text)

        if text in _translation_cache:
            continue

        translated = ""
        try:
            result = _GoogleTranslator(source="th", target="en").translate(text)
            translated = _clean_translation_output(str(result or text).strip())
        except Exception:
            translated = _translate_with_argos(text)

        if translated:
            with _translation_lock:
                _translation_cache[text] = translated
                _translation_cache_dirty = True
            completed += 1

        if completed > 0 and completed % _BG_TRANSLATE_CLEAR_EVERY == 0:
            # Persist progress, but do NOT clear the payload cache mid-backlog —
            # that wiped the cache between requests and forced a ~46s rebuild every
            # time. Translations are applied once the queue fully drains (below).
            _save_translation_cache()

    _save_translation_cache()
    # NOTE: deliberately do NOT clear the payload cache here. Clearing it forced a
    # ~46s rebuild on the next request and, while a backlog drained, kept the
    # dashboard slow. New translations persist to disk and are picked up on the
    # next genuine cache miss (data upload / refresh / restart). Stability and
    # speed are prioritised over surfacing translations mid-session.


def _queue_background_translation(text: str) -> None:
    """Add text to the background translation queue and start the worker if idle."""
    global _translation_worker_active
    if not _DEEP_TRANSLATOR_AVAILABLE:
        return
    with _translation_lock:
        _translation_pending.add(text)
        if _translation_worker_active:
            return
        _translation_worker_active = True
    t = threading.Thread(target=_bg_translation_worker, daemon=True, name="translate-bg")
    t.start()


def _local_dict_translate(text: str) -> str:
    """Best-effort partial translation using the local dictionary (instant, offline)."""
    translated = text
    for thai, english in sorted(THAI_MAINTENANCE_TRANSLATIONS.items(), key=lambda item: len(item[0]), reverse=True):
        translated = translated.replace(thai, f" {english} ")
    translated = re.sub(r"\bAir\s*blast\s*(\d+)", r"Air Blast \1", translated, flags=re.IGNORECASE)
    translated = re.sub(r"\bFreezer\b", "freezer", translated, flags=re.IGNORECASE)
    translated = re.sub(r"\bChiller\b", "chiller", translated, flags=re.IGNORECASE)
    translated = re.sub(r"\bCold\b", "cold room / cold area", translated, flags=re.IGNORECASE)
    return _clean_translation_output(translated)


def translate_maintenance_description(value):
    original = clean_unicode_text(value)
    if not original:
        return ""
    if not has_thai_text(original):
        return original
    _refresh_translation_cache_if_changed()
    # Return cached translation instantly.
    cached = _translation_cache.get(original)
    if cached is not None:
        return _clean_translation_output(cached)
    # Not yet cached: queue for background Google Translate and return local-dict result now.
    _queue_background_translation(original)
    return _local_dict_translate(original)


THAI_MAINTENANCE_TRANSLATIONS = {
    "แรงดันน้ำยาต่ำ": "refrigerant pressure low",
    "แรงดันน้ำยา": "refrigerant pressure",
    "ตรวจเช็ค": "inspect / check",
    "น้ำรั่ว": "water leaking",
    "ไม่ทำงาน": "not working",
    "ไม่ติด": "not turning on / not lighting",
    "หลอดไฟ": "light bulb",
    "สายยาง": "hose",
    "ก๊อกน้ำ": "water tap",
    "ก็อกน้ำ": "water tap",
    "น้ำยา": "refrigerant",
    "เช็ค": "check",
    "แก้ไข": "repair / fix",
    "เปลี่ยน": "replace",
    "ซ่อม": "repair",
    "พัง": "broken / faulty",
    "เสีย": "faulty",
    "รั่ว": "leaking",
    "ห้อง": "room",
    "ประตู": "door",
    "ไฟ": "light / electricity",
    "ตู้": "cabinet / unit",
    "คอมเพรสเซอร์": "compressor",
    "คอม": "compressor",
    "คอยล์": "coil",
    "พัดลม": "fan",
    "มอเตอร์": "motor",
    "ปั๊ม": "pump",
    "วาล์ว": "valve",
    "ท่อ": "pipe",
    "สตีม": "steam",
    "หลุด": "loose / disconnected",
    "ตก": "fallen / dropped",
    "ราง": "rail",
    "ฝั่งดิบ": "raw side",
    "ฝั่งสุก": "cooked side",
    "ทางไป": "to",
    "จุด": "points",
}


def clean_job_trade(value):
    cleaned = clean_ascii_text(value, "Work Order")
    normalized = normalize_key(cleaned)
    mapping = {
        "building": "Building",
        "aircondition": "Air Condition",
        "electricalsystem": "Electrical System",
        "coolingmachine": "Cooling Machine",
        "productionmachine": "Production Machine",
        "toolandequipment": "Tool and Equipment",
        "watersystem": "Water System",
        "utilitymachine": "Utility Machine",
    }
    return mapping.get(normalized, cleaned.title() if cleaned.islower() else cleaned)


def parse_ttr_hours(value):
    if value is None or pd.isna(value):
        return None
    text = str(value).strip().lower()
    if not text:
        return None
    is_negative = bool(re.match(r"^-\s*\d", text))

    numeric = pd.to_numeric(text, errors="coerce")
    if pd.notna(numeric):
        return float(numeric)

    hours = 0.0
    minutes = 0.0
    hour_match = re.search(r"([\d.]+)\s*(?:hr|hour|hours|h)\b", text)
    minute_match = re.search(r"([\d.]+)\s*(?:min|minute|minutes|m)\b", text)
    if hour_match:
        hours = float(hour_match.group(1))
    if minute_match:
        minutes = float(minute_match.group(1))
    if not hour_match and not minute_match:
        compact = re.findall(r"[\d.]+", text)
        if compact:
            hours = float(compact[0])

    total = hours + (minutes / 60)
    if is_negative:
        return -abs(total) if total else None
    return total if total > 0 else None


def parse_work_order_date_parts(row, suffix="", default_year=None):
    year_value = row.get(f"Year{suffix}")
    month_value = row.get(f"Month{suffix}")
    day_value = row.get(f"Day{suffix}")

    year = pd.to_numeric(year_value, errors="coerce")
    day = pd.to_numeric(day_value, errors="coerce")
    if pd.isna(month_value) or pd.isna(day):
        return None
    if pd.isna(year):
        if default_year is None:
            return None
        year = default_year

    parsed = pd.to_datetime(
        f"{int(day)} {str(month_value).strip()} {int(year)}",
        errors="coerce",
        dayfirst=True,
    )
    if pd.isna(parsed):
        return None
    return parsed.to_pydatetime().replace(hour=0, minute=0, second=0, microsecond=0)


def infer_risk_area(*values):
    haystack = " ".join(str(value or "") for value in values).lower()
    if "high risk" in haystack:
        return "High Risk"
    if "medium risk" in haystack or "มีเดียม" in haystack:
        return "Medium Risk"
    if "low risk" in haystack:
        return "Low Risk"
    return None


def extract_english_description_name(description):
    cleaned = clean_ascii_text(description)
    if not cleaned:
        return ""

    stop_words = {
        "fix", "repair", "install", "replace", "change", "check", "broken", "leak",
        "abnormal", "not", "working", "work", "area", "side", "room",
    }
    tokens = re.findall(r"[A-Za-z][A-Za-z0-9-]*|\d+", cleaned)
    useful = [token for token in tokens if token.lower() not in stop_words]
    if not useful:
        return ""
    return " ".join(useful[:5])



def get_path_signature(path):
    try:
        stat = os.stat(path)
    except (FileNotFoundError, OSError):
        return None
    return (stat.st_mtime_ns, stat.st_size)


def percentile(values, fraction):
    if not values:
        return None
    ordered = sorted(values)
    index = int(round((len(ordered) - 1) * fraction))
    return ordered[max(0, min(index, len(ordered) - 1))]


def format_hours(hours):
    if hours is None:
        return None
    if hours < 1:
        return f"{round(hours * 60)} min"
    if hours >= 24:
        days = max(1, round(hours / 24))
        return f"{days} {'day' if days == 1 else 'days'}"

    whole_hours = int(hours)
    minutes = round((hours - whole_hours) * 60)
    if minutes == 60:
        return f"{whole_hours + 1} hr"
    if minutes > 0:
        return f"{whole_hours} hr {minutes} min"
    return f"{whole_hours} hr"


def load_work_order_downtime():
    sources = get_work_order_source_paths()
    # Asset mapping edits must invalidate the enriched work-order cache too,
    # otherwise criticality/group changes can stay stale until a restart.
    sig = (
        ("asset_master", get_path_signature(get_asset_master_path(DATA_DIR))),
        *tuple((p, get_path_signature(p)) for p in sources),
    )
    if _WO_LOAD_CACHE["sig"] == sig and _WO_LOAD_CACHE["payload"] is not None:
        return _WO_LOAD_CACHE["payload"]

    records = []
    source_paths = []
    read_errors = []

    work_order_sources = sources

    for exported_path in (path for path in work_order_sources if os.path.exists(path)):
        source_paths.append(exported_path)
        try:
            df = read_work_order_source_file(exported_path)
            file_end_time = datetime.fromtimestamp(os.path.getmtime(exported_path))
            default_year = infer_default_year_from_path(exported_path)
            has_start_date = {"Month", "Day"}.issubset(set(df.columns))
            has_end_date = {"Month.1", "Day.1"}.issubset(set(df.columns))
            has_request_created_column = "Request Created Date" in df.columns

            for _, row in df.iterrows():
                raw_ttr_value = get_first_present(row, "TTR(hr)", "TTR", "downtime_hours")
                imported_ttr_hours = parse_ttr_hours(raw_ttr_value)
                if imported_ttr_hours is None:
                    ttr_minutes = pd.to_numeric(get_first_present(row, "TTR Minutes", "TTR(min)", "TTR Minutes "), errors="coerce")
                    if pd.notna(ttr_minutes) and float(ttr_minutes) > 0:
                        imported_ttr_hours = float(ttr_minutes) / 60

                request_created_raw = get_first_present(row, "Request Created Date")
                if not has_request_created_column:
                    request_created_raw = get_first_present(row, "Actual Start Date", "Maintenance Start Date", "Start Date")
                request_created_time, request_created_invalid = parse_work_order_datetime_with_quality(request_created_raw)

                actual_start_raw = get_first_present(row, "Actual Start Date", "Maintenance Start Date", "Start Date")
                actual_start_time, actual_start_invalid = parse_work_order_datetime_with_quality(actual_start_raw)
                if actual_start_time is None and has_start_date:
                    actual_start_time = parse_work_order_date_parts(row, "", default_year=default_year)
                    actual_start_invalid = False

                actual_end_raw = get_first_present(row, "Actual End Date", "Maintenance End Date", "End Date", "Closed Date", "Completed Date")
                actual_end_time, actual_end_invalid = parse_work_order_datetime_with_quality(actual_end_raw, end_of_day=True)
                if actual_end_time is None and has_end_date:
                    actual_end_time = parse_work_order_date_parts(row, ".1", default_year=default_year)
                    if actual_end_time is not None:
                        actual_end_time = actual_end_time.replace(hour=23, minute=59, second=59)
                    actual_end_invalid = False

                raw_machine_id = get_first_present(row, "Machine ID", "machine_code", "Asset ID", "AssetID") or ""
                raw_machine_name = get_first_present(row, "Machine Name", "machine_name", "Asset Name", "Equipment Name") or ""
                description = get_first_present(row, "Description", "Notes", "Remarks") or ""
                risk_area = infer_risk_area(raw_machine_name, description)
                job_trade = clean_job_trade(get_first_present(row, "JobTrade", "Job Trade", "system", "System"))
                maintenance_job_type = clean_unicode_text(get_first_present(row, "JobTypeId", "Maintenance job type", "Maintenance job type variant", "Maintenance request type"))
                raw_location_value = get_first_present(row, "Location", "Building", "Area", "Functional location")
                description_original = clean_unicode_text(description)
                translated_description = translate_maintenance_description(description_original)
                machine_equipment_name = clean_unicode_text(raw_machine_name)

                machine_code = clean_ascii_text(raw_machine_id, "WO-Asset")
                clean_machine_name = clean_ascii_text(raw_machine_name)
                if risk_area:
                    machine_name = f"{risk_area} Production Area"
                elif clean_machine_name and not has_thai_text(raw_machine_name):
                    machine_name = clean_machine_name
                else:
                    machine_name = extract_english_description_name(description) or machine_code or "Work Order Asset"
                default_area = clean_ascii_text(get_first_present(row, "Location", "Building", "Area"), "Work Area")

                area = risk_area or default_area or "Unassigned"
                status = clean_ascii_text(get_first_present(row, "Request State", "Status", "request_state"), "Review")
                work_order_id = clean_ascii_text(get_first_present(row, "WO ID", "work_order_id", "Work Order ID"))
                maintenance_request_id = clean_ascii_text(get_first_present(row, "Request ID", "maintenance_order_id", "Request No"))
                acknowledgement_status = calculate_acknowledgement_status(status, work_order_id)
                data_quality_flags = build_work_order_quality_flags(
                    status,
                    request_created_raw,
                    request_created_time,
                    request_created_invalid,
                    actual_start_raw,
                    actual_start_time,
                    actual_start_invalid,
                    actual_end_raw,
                    actual_end_time,
                    actual_end_invalid,
                )
                valid_finished_record = (
                    data_quality_flags == ["Valid"]
                    and is_finished_lifecycle(status)
                    and request_created_time is not None
                    and actual_start_time is not None
                    and actual_end_time is not None
                    and actual_end_time >= actual_start_time
                    and actual_end_time >= request_created_time
                )

                raw_priority = get_first_present(row, "Priority", "priority", "Priority No", "Priority Number")
                try:
                    priority = int(float(str(raw_priority).strip())) if raw_priority is not None else None
                    if priority is not None and not (1 <= priority <= 10):
                        priority = None
                except (ValueError, TypeError):
                    priority = None

                downtime_hours = None
                ttr_source = "excluded_status"
                if valid_finished_record:
                    if imported_ttr_hours is not None and imported_ttr_hours > 0:
                        downtime_hours = imported_ttr_hours
                        ttr_source = "imported_ttr"
                    else:
                        downtime_hours = (actual_end_time - actual_start_time).total_seconds() / 3600
                        ttr_source = "date_derived"
                elif is_finished_lifecycle(status):
                    ttr_source = "invalid_finished_dates"

                # Created date drives MR demand and open-load selection; Actual dates drive TTR/MTBF only for valid Finished MR.
                downtime_start_time = actual_start_time if valid_finished_record else request_created_time
                downtime_end_time = actual_end_time if valid_finished_record else None
                maintenance_start_time = actual_start_time
                maintenance_end_time = actual_end_time

                records.append(
                    {
                        "system": job_trade,
                        "machine_code": machine_code,
                        "machine_name": machine_name,
                        "area": area,
                        "source": "Work Order",
                        "status": status,
                        "detection_type": "Work Order",
                        "start_time": downtime_start_time.isoformat() if downtime_start_time else None,
                        "end_time": downtime_end_time.isoformat() if downtime_end_time else None,
                        "maintenance_start_time": maintenance_start_time.isoformat() if maintenance_start_time else None,
                        "maintenance_end_time": maintenance_end_time.isoformat() if maintenance_end_time else None,
                        "request_created_time": request_created_time.isoformat() if request_created_time else None,
                        "timing_context": "Created date used for MR demand/open load; Actual start/end used for valid Finished MR",
                        "duration_context": "Maintenance resolution time from work-order TTR" if ttr_source == "imported_ttr" else ("Maintenance resolution time derived from valid Finished start/end dates" if ttr_source == "date_derived" else "Excluded from MTTR/TTR by lifecycle or data-quality rule"),
                        "duration_hours": round(float(downtime_hours), 3) if downtime_hours is not None else None,
                        "ttr_source": ttr_source or "missing",
                        "raw_ttr": raw_ttr_value,
                        "is_critical": area == "High Risk",
                        "priority": priority,
                        "service_level": clean_ascii_text(raw_priority),
                        "work_order_id": work_order_id,
                        "maintenance_order_id": maintenance_request_id,
                        "started_by": clean_ascii_text(get_first_present(row, "Started By", "started_by")),
                        "created_by": clean_ascii_text(get_first_present(row, "Created By", "created_by")),
                        "machine_equipment_name": machine_equipment_name,
                        "description_original": description_original,
                        "translated_description": translated_description,
                        "maintenance_job_type": maintenance_job_type,
                        "raw_functional_location": clean_unicode_text(raw_location_value),
                        "acknowledgement_status": acknowledgement_status,
                        "data_quality_flag": "Valid" if data_quality_flags == ["Valid"] else "; ".join(data_quality_flags),
                        "data_quality_flags": data_quality_flags,
                        "valid_mttr_ttr": valid_finished_record,
                        "status_category": "Open" if is_new_lifecycle(status) or is_in_progress_lifecycle(status) else ("Closed" if is_finished_lifecycle(status) else "Review"),
                        "remarks": description_original,
                        "raw_machine_id": clean_ascii_text(raw_machine_id),
                        "raw_machine_name": machine_equipment_name,
                        "raw_location": clean_unicode_text(raw_location_value),
                        "source_path": exported_path,
                    }
                )
        except Exception as exc:
            read_errors.append(str(exc))

    if not source_paths:
        return {
            "available": False,
            "records": [],
            "message": "No work order downtime source connected yet.",
            "last_synced": None,
        }

    latest_synced = max(datetime.fromtimestamp(os.path.getmtime(path)) for path in source_paths)
    if not records and read_errors:
        return {
            "available": False,
            "records": [],
            "message": "Work order downtime source could not be read.",
            "last_synced": latest_synced.isoformat(),
        }

    records = enrich_work_order_records(records, DATA_DIR)

    # Annotate every record with the authoritative resolved_stage so all downstream
    # calculations read the same value without re-running resolution logic.
    # resolve_work_order_stage() combines text-based detection with asset mapping —
    # see its docstring for priority rules.
    for rec in records:
        rec["resolved_stage"] = resolve_work_order_stage(rec)

    result = {
        "available": True,
        "records": records,
        "message": "Work order downtime source loaded.",
        "last_synced": latest_synced.isoformat(),
    }
    _WO_LOAD_CACHE["sig"] = sig
    _WO_LOAD_CACHE["payload"] = result
    return result


def within_period(iso_string, start_dt, end_dt):
    dt = parse_iso_datetime(iso_string)
    if dt is None:
        return False
    return start_dt <= dt <= end_dt


def overlaps_period(start_iso, end_iso, start_dt, end_dt):
    start_value = parse_iso_datetime(start_iso)
    end_value = parse_iso_datetime(end_iso)
    if start_value is None and end_value is None:
        return False
    if start_value is None:
        start_value = end_value
    if end_value is None:
        end_value = start_value
    return start_value <= end_dt and end_value >= start_dt



def build_asset_breakdown(events):
    grouped = {}
    for event in events:
        key = event["machine_code"]
        row = grouped.setdefault(
            key,
            {
                "machine_code": event["machine_code"],
                "machine_name": event["machine_name"],
                "system": event["system"],
                "area": event["area"],
                "downtime_hours": 0.0,
                "event_count": 0,
                "is_critical": bool(event.get("is_critical")),
            },
        )
        row["downtime_hours"] += float(event.get("duration_hours") or 0)
        row["event_count"] += 1

    rows = sorted(grouped.values(), key=lambda item: (-item["downtime_hours"], -item["event_count"], item["machine_name"]))
    for row in rows:
        row["downtime_hours"] = round(row["downtime_hours"], 3)
    return rows


def build_breakdown_rows(events, key_name):
    grouped = {}
    for event in events:
        label = event.get(key_name) or "Unassigned"
        row = grouped.setdefault(label, {"label": label, "downtime_hours": 0.0, "event_count": 0})
        row["downtime_hours"] += float(event.get("duration_hours") or 0)
        row["event_count"] += 1

    rows = sorted(grouped.values(), key=lambda item: (-item["downtime_hours"], -item["event_count"], item["label"]))
    for row in rows:
        row["downtime_hours"] = round(row["downtime_hours"], 3)
    return rows


def build_trend_series(events, start_dt, end_dt):
    labels = []
    hours = []
    counts = []

    current = start_dt.replace(hour=0, minute=0, second=0, microsecond=0)
    end_date = end_dt.replace(hour=0, minute=0, second=0, microsecond=0)

    while current <= end_date:
        next_day = current + timedelta(days=1)
        day_events = [event for event in events if within_period(event.get("start_time"), current, next_day)]
        labels.append(current.strftime("%d %b"))
        hours.append(round(sum(event["duration_hours"] for event in day_events), 3))
        counts.append(len(day_events))
        current = next_day

    return {"labels": labels, "downtime_hours": hours, "event_counts": counts}


def build_cache_signature(period, month_filter=None, start_filter=None, end_filter=None, work_orders_only=False, stage_filter=None, allow_excel_fallback=True):
    signatures = [
        DOWNTIME_CACHE_VERSION,
        period,
        month_filter,
        start_filter,
        end_filter,
        normalize_stage_filter(stage_filter),
        bool(allow_excel_fallback),
    ]
    signatures.append(get_path_signature(get_asset_master_path(DATA_DIR)))
    for source_path in get_work_order_source_paths():
        signatures.append((source_path, get_path_signature(source_path)))
    return tuple(signatures)


def build_downtime_payload(period=None, month=None, start=None, end=None, work_orders_only=False, stage=None, allow_excel_fallback=True):
    normalized_period = normalize_period(period)
    normalized_month = normalize_month_filter(month)
    normalized_stage = normalize_stage_filter(stage)
    custom_start = normalize_date_filter(start)
    custom_end = normalize_date_filter(end)
    custom_start_key = custom_start.strftime("%Y-%m-%d") if custom_start else None
    custom_end_key = custom_end.strftime("%Y-%m-%d") if custom_end else None
    cache_signature = build_cache_signature(
        normalized_period,
        normalized_month,
        custom_start_key,
        custom_end_key,
        work_orders_only,
        normalized_stage,
        allow_excel_fallback,
    )
    cached = _DOWNTIME_CACHE.get(cache_signature)
    if cached is not None:
        return cached

    sql_mapping_meta = _sql_asset_mapping_meta()

    # Phase 3: prefer SQL over re-reading Excel. Fall back to Excel only when allowed.
    if _sql_has_work_orders():
        work_order_payload = load_work_order_downtime_sql(normalized_stage)
        work_order_events = list(work_order_payload.get("records") or [])
    elif allow_excel_fallback:
        work_order_payload = load_work_order_downtime()
        work_order_events = filter_work_orders_by_stage(
            work_order_payload.get("records") or [], normalized_stage
        )
    else:
        work_order_payload = {
            "available": False,
            "records": [],
            "message": "No work order SQL data found for SQL-only overview loading.",
            "last_synced": None,
        }
        work_order_events = []
    sla_target_config = load_sla_target_config(DATA_DIR) if allow_excel_fallback else _default_sla_target_config()
    latest_timestamps = []
    if work_order_payload.get("last_synced"):
        latest_timestamps.append(pd.Timestamp(work_order_payload["last_synced"]))

    reference_dt = max(latest_timestamps) if latest_timestamps else None
    if reference_dt is None:
        payload = {
            "meta": {
                "period": normalized_period,
                "period_label": get_period_label(normalized_period),
                "month": normalized_month,
                "month_label": format_month_label(normalized_month) if normalized_month else "All Months",
                "reference_end": None,
                "last_synced": None,
                "work_order_available": work_order_payload["available"],
                "stage_filter": normalized_stage,
                "stage_label": normalized_stage or "All Stages",
            },
            "summary": {
                "total_hours": None,
                "event_count": 0,
                "avg_event_hours": None,
                "longest_event_hours": None,
                "highest_system": None,
                "work_order_hours": None,
                "work_order_record_count": 0,
            },
            "alerts": [],
            "trend": {"labels": [], "downtime_hours": [], "event_counts": []},
            "system_breakdown": [],
            "source_breakdown": [
                {
                    "label": "Work Order",
                    "downtime_hours": None,
                    "available": work_order_payload["available"],
                    "message": work_order_payload["message"],
                },
            ],
            "area_breakdown": [],
            "asset_breakdown": [],
            "events": [],
            "filters": {"systems": [], "areas": [], "sources": ["Work Order"], "stages": STAGE_FILTER_OPTIONS},
            "months": [],
            "work_order_source": work_order_payload,
            "config": {
                "sla_targets": sla_target_config,
            },
            "operating_windows": [],
            "management": {
                "summary": {
                    "total_downtime_hours": 0.0,
                    "total_work_orders": 0,
                    "overall_mttr_hours": None,
                    "critical_downtime_hours": 0.0,
                    "non_critical_facility_downtime_hours": 0.0,
                    "open_work_orders": 0,
                    "highest_mttr_machine_group": None,
                    "highest_mttr_hours": None,
                    "most_affected_location": None,
                    "most_affected_location_hours": None,
                    "critical_machine_groups_with_repeats": 0,
                },
                "criticality_rows": [],
                "machine_group_rows": [],
                "location_rows": [],
                "trend": {"labels": [], "downtime_hours": [], "work_order_counts": [], "bucket_mode": "day"},
                "work_orders": [],
                "filters": {"criticalities": [], "machine_groups": [], "locations": [], "asset_ids": [], "statuses": []},
                "alerts": [],
                "mapping_meta": sql_mapping_meta if not allow_excel_fallback else get_grouped_machine_mapping_meta(DATA_DIR),
            },
        }
        _DOWNTIME_CACHE[cache_signature] = payload
        return payload

    month_options = list(reversed(build_year_month_options_from_events(work_order_events, reference_dt)))

    if normalized_period == "custom":
        if custom_start and custom_end:
            period_start = custom_start
            period_end = custom_end.replace(hour=23, minute=59, second=59, microsecond=999999)
            if period_end < period_start:
                first = period_end.replace(hour=0, minute=0, second=0, microsecond=0)
                second = period_start.replace(hour=23, minute=59, second=59, microsecond=999999)
                period_start, period_end = first, second
        else:
            normalized_period = "ytd"
            period_end = reference_dt
            period_start = reference_dt.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    elif normalized_period == "this_month":
        target_month = normalized_month or reference_dt.strftime("%Y-%m")
        month_start = pd.to_datetime(f"{target_month}-01").to_pydatetime().replace(hour=0, minute=0, second=0, microsecond=0)
        if month_start.year == reference_dt.year and month_start.month == reference_dt.month:
            period_end = reference_dt
        elif month_start.month == 12:
            period_end = month_start.replace(year=month_start.year + 1, month=1) - timedelta(microseconds=1)
        else:
            period_end = month_start.replace(month=month_start.month + 1) - timedelta(microseconds=1)
        period_start = month_start
        normalized_month = target_month
    elif normalized_period == "last_month":
        current_month_start = reference_dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if current_month_start.month == 1:
            period_start = current_month_start.replace(year=current_month_start.year - 1, month=12)
        else:
            period_start = current_month_start.replace(month=current_month_start.month - 1)
        period_end = current_month_start - timedelta(microseconds=1)
    elif normalized_month:
        month_start = pd.to_datetime(f"{normalized_month}-01").to_pydatetime().replace(hour=0, minute=0, second=0, microsecond=0)
        if month_start.month == 12:
            period_end = month_start.replace(year=month_start.year + 1, month=1) - timedelta(microseconds=1)
        else:
            period_end = month_start.replace(month=month_start.month + 1) - timedelta(microseconds=1)
        period_start = month_start
    elif normalized_period == "ytd":
        period_end = reference_dt
        period_start = reference_dt.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    elif normalized_period == "previous_year":
        previous_year = reference_dt.year - 1
        period_start = datetime(previous_year, 1, 1)
        period_end = datetime(previous_year, 12, 31, 23, 59, 59, 999999)
    elif normalized_period == "all_years":
        all_timestamps = []
        for event in work_order_events:
            for key in ("start_time", "end_time"):
                parsed = parse_iso_datetime(event.get(key))
                if parsed is not None:
                    all_timestamps.append(parsed)
                    break
        period_start = min(all_timestamps).replace(hour=0, minute=0, second=0, microsecond=0) if all_timestamps else reference_dt
        period_end = reference_dt
    else:
        period_days = get_period_days(normalized_period)
        period_end = reference_dt
        period_start = (reference_dt - timedelta(days=period_days - 1)).replace(hour=0, minute=0, second=0, microsecond=0)

    selected_work_orders = (
        list(work_order_events)
        if normalized_period == "all_years"
        else [event for event in work_order_events if overlaps_period(event.get("start_time"), event.get("end_time"), period_start, period_end)]
    )
    selected_valid_work_orders = [event for event in selected_work_orders if event.get("duration_hours") is not None]
    selected_events = selected_valid_work_orders

    management_payload = build_management_downtime_payload(
        selected_work_orders,
        [],
        period_start,
        period_end,
        DATA_DIR,
        mtbf_records=selected_work_orders,
        historical_records=work_order_events,
        mapping_meta=sql_mapping_meta if not allow_excel_fallback else None,
    )

    total_hours = round(sum(float(event.get("duration_hours") or 0) for event in selected_events), 3) if selected_events else 0.0
    event_count = len(selected_events)
    longest_event_hours = max((float(event.get("duration_hours") or 0) for event in selected_events), default=None)
    avg_event_hours = round(total_hours / event_count, 3) if event_count else None

    system_breakdown = build_breakdown_rows(selected_events, "system")
    area_breakdown = build_breakdown_rows(selected_events, "area")
    asset_breakdown = build_asset_breakdown(selected_events)
    highest_system = system_breakdown[0]["label"] if system_breakdown else None

    work_order_hours = round(total_hours, 3) if work_order_payload["available"] else None

    alerts = []
    if highest_system and system_breakdown and system_breakdown[0]["downtime_hours"] >= 3:
        alerts.append({
            "level": "warning",
            "message": f"{highest_system} has the highest TTR logged in the selected period ({format_hours(system_breakdown[0]['downtime_hours'])}).",
        })
    if longest_event_hours and longest_event_hours >= 2:
        longest_event = max(selected_events, key=lambda event: event["duration_hours"])
        machine_name = longest_event.get("machine_name") or longest_event.get("machine_code") or "Unknown"
        alerts.append({
            "level": "critical",
            "message": f"{machine_name} recorded the longest TTR ({format_hours(longest_event_hours)}).",
        })

    filters = {
        "systems": sorted({event["system"] for event in selected_events}),
        "areas": sorted({event["area"] for event in selected_events}),
        "sources": ["Work Order"] if work_order_payload["available"] else [],
        "stages": STAGE_FILTER_OPTIONS,
    }

    payload = {
        "meta": {
            "period": normalized_period,
            "period_label": get_period_label(normalized_period),
            "month": normalized_month,
            "month_label": format_month_label(normalized_month) if normalized_month else "All Months",
            "period_start": period_start.isoformat(),
            "period_end": period_end.isoformat(),
            "custom_start": custom_start_key,
            "custom_end": custom_end_key,
            "reference_end": period_end.isoformat(),
            "last_synced": max(latest_timestamps).isoformat(),
            "work_order_available": work_order_payload["available"],
            "work_orders_only": True,
            "stage_filter": normalized_stage,
            "stage_label": normalized_stage or "All Stages",
            "all_years_warning": (
                "All Years view is for long-term reliability analysis."
                if normalized_period == "all_years"
                else None
            ),
        },
        "summary": {
            "total_hours": total_hours,
            "event_count": event_count,
            "avg_event_hours": avg_event_hours,
            "longest_event_hours": round(longest_event_hours, 3) if longest_event_hours is not None else None,
            "highest_system": highest_system,
            "work_order_hours": work_order_hours,
            "work_order_record_count": len(selected_work_orders),
        },
        "alerts": alerts,
        "trend": build_trend_series(selected_events, period_start, period_end),
        "system_breakdown": system_breakdown,
        "source_breakdown": [{
            "label": "Work Order",
            "downtime_hours": work_order_hours,
            "available": work_order_payload["available"],
            "message": "Imported work order TTR logged as maintenance resolution time.",
        }],
        "area_breakdown": area_breakdown,
        "asset_breakdown": asset_breakdown[:8],
        "events": sorted(selected_events, key=lambda event: (event.get("start_time") or "", event.get("machine_name") or ""), reverse=True),
        "filters": filters,
        "months": [{"value": value, "label": format_month_label(value)} for value in month_options],
        "work_order_source": work_order_payload,
        "config": {
            "sla_targets": sla_target_config,
        },
        "operating_windows": [],
        "management": management_payload,
    }

    _DOWNTIME_CACHE[cache_signature] = payload
    return payload


def build_downtime_cache_document():
    default_payload = build_downtime_payload("ytd")
    month_values = [row["value"] for row in default_payload.get("months", [])]

    payloads = {}
    for period in ("this_month", "last_month", "ytd", "last12", "previous_year", "all_years"):
        payloads[period] = build_downtime_payload(period)

    for month_value in month_values:
        payloads[f"this_month:{month_value}"] = build_downtime_payload("this_month", month_value)

    return {
        "generated_at": datetime.now().isoformat(),
        "months": month_values,
        "payloads": payloads,
    }


def write_downtime_cache_file(output_path=None):
    target_path = output_path or DOWNTIME_CACHE_OUTPUT_FILE
    os.makedirs(os.path.dirname(target_path), exist_ok=True)
    document = build_downtime_cache_document()
    with open(target_path, "w", encoding="utf-8") as handle:
        json.dump(document, handle, ensure_ascii=False)
    return target_path
